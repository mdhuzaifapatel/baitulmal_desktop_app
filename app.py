from fastapi import FastAPI, Request, Form, HTTPException
from fastapi.responses import RedirectResponse, HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, and_
from datetime import datetime, date, timedelta
import shutil
import os
from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import io
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
import tkinter as tk
from tkinter import filedialog
import sys

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def prompt_save_path(title, default_name, file_types):
    try:
        root = tk.Tk()
        root.attributes("-topmost", True)
        root.withdraw()
        filepath = filedialog.asksaveasfilename(
            title=title,
            initialfile=default_name,
            filetypes=file_types
        )
        root.destroy()
        return filepath
    except Exception:
        return None

def prompt_open_path(title, file_types):
    try:
        root = tk.Tk()
        root.attributes("-topmost", True)
        root.withdraw()
        filepath = filedialog.askopenfilename(
            title=title,
            filetypes=file_types
        )
        root.destroy()
        return filepath
    except Exception:
        return None
from database import SessionLocal, engine
from models import *
from seed_data import *

# Create tables
Base.metadata.create_all(bind=engine)

app = FastAPI(title="Baitulmal Donation Management System")

# Add session middleware for authentication
app.add_middleware(SessionMiddleware, secret_key="baitulmal_secret_key_2024", max_age=3600)

# Default admin credentials (plain text)
DEFAULT_ADMIN_USERNAME = "admin"
DEFAULT_ADMIN_PASSWORD = "admin"

# Direct Jinja2 environment
env = Environment(loader=FileSystemLoader(resource_path("templates")), autoescape=True)

# -----------------------------------------------------------------------
# Expense Category Classification Map
# Maps every known expense category name to one of three classes:
#   'charity'   – Direct welfare / charitable aid to beneficiaries
#   'asset'     – Capital expenditure / fixed asset additions
#   'admin'     – Administrative / operational NGO overhead
# -----------------------------------------------------------------------
EXPENSE_CLASSIFICATION = {
    # Direct Charity & Welfare
    "ambulance service":         "charity",
    "imdad to needy":            "charity",
    "mass marriages":            "charity",
    "monthly pension for needy": "charity",
    "ration kit":                "charity",
    "staff welfare expenses":    "charity",
    # Capital / Fixed Assets
    "fixed assets":              "asset",
    # Administrative / Operational Overhead
    "advertisement expenses":    "admin",
    "bank charges":              "admin",
    "fuel expenses":             "admin",
    "general expenditure":       "admin",
    "miscellaneous":             "admin",
    "printing and stationary":   "admin",
    "repair and maintenance":    "admin",
    "staff salary":              "admin",
    "travel expenses":           "admin",
    "vehicle service":           "admin",
    "wages":                     "admin",
}

def classify_expense(name: str) -> str:
    """Return the classification for a category name (case-insensitive)."""
    return EXPENSE_CLASSIFICATION.get(name.strip().lower(), "charity")  # default: charity


def get_financial_year_range(fy_year: int):
    """
    Get the date range for a given financial year.
    Financial Year: April 1 to March 31
    Example: FY 2024-2025 means April 1, 2024 to March 31, 2025
    """
    start_date = date(fy_year, 4, 1)
    end_date = date(fy_year + 1, 3, 31)
    return start_date, end_date

def get_current_financial_year():
    """
    Get the current financial year (returns the starting year).
    Example: If current date is May 2024, returns 2024
    If current date is February 2025, returns 2024
    """
    today = date.today()
    if today.month >= 4:
        return today.year
    else:
        return today.year - 1

def get_date_range_for_filter(filter_type: str, start_date_str: str = None, end_date_str: str = None, fy: str = None):
    if fy is not None:
        try:
            fy = int(str(fy).split('?')[0].split('&')[0])
        except ValueError:
            fy = None
    if filter_type == "current_month":
        today = date.today()
        actual_start_date = date(today.year, today.month, 1)
        if today.month == 12:
            actual_end_date = date(today.year + 1, 1, 1) - timedelta(days=1)
        else:
            actual_end_date = date(today.year, today.month + 1, 1) - timedelta(days=1)
        title_suffix = f"{today.strftime('%B %Y')}"
    elif filter_type == "manual" and start_date_str and end_date_str:
        actual_start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        actual_end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        title_suffix = f"{actual_start_date.strftime('%d-%m-%Y')} to {actual_end_date.strftime('%d-%m-%Y')}"
    else:
        # Default to current FY
        if fy is None:
            fy = get_current_financial_year()
        actual_start_date, actual_end_date = get_financial_year_range(fy)
        title_suffix = f"FY {fy}-{fy+1}"
    
    return actual_start_date, actual_end_date, title_suffix

def get_available_financial_years(db: Session):
    """
    Get list of financial years that either have receipts or is the current FY.
    """
    from sqlalchemy import text
    current_fy = get_current_financial_year()
    
    # SQL to get distinct FYs from receipts in SQLite
    sql = text("""
        SELECT DISTINCT 
            CASE WHEN CAST(strftime('%m', date) AS INTEGER) >= 4 
                 THEN CAST(strftime('%Y', date) AS INTEGER) 
                 ELSE CAST(strftime('%Y', date) AS INTEGER) - 1 
            END as fy 
        FROM receipts
    """)
    
    results = db.execute(sql).fetchall()
    fys = {current_fy}
    for row in results:
        if row[0] is not None:
            fys.add(int(row[0]))
            
    return sorted(list(fys))

def get_balances(db, start_date, end_date):
    """Calculate opening, period, and closing balances."""
    # Sum of all initial balances (global)
    initial_total = db.query(func.sum(PaymentMode.initial_balance)).scalar() or 0
    
    # Opening Receipts
    opening_receipts = db.query(func.sum(Receipt.amount)).filter(Receipt.date < start_date).scalar() or 0
    # Opening Expenses
    opening_expenses = db.query(func.sum(Expense.amount)).filter(Expense.date < start_date).scalar() or 0
    
    opening_balance = initial_total + opening_receipts - opening_expenses
    
    # Period Receipts
    period_receipts = db.query(func.sum(Receipt.amount)).filter(
        and_(Receipt.date >= start_date, Receipt.date <= end_date)
    ).scalar() or 0
    # Period Expenses
    period_expenses = db.query(func.sum(Expense.amount)).filter(
        and_(Expense.date >= start_date, Expense.date <= end_date)
    ).scalar() or 0
    
    closing_balance = opening_balance + period_receipts - period_expenses
    return opening_balance, period_receipts, closing_balance, period_expenses

# Static files
app.mount("/static", StaticFiles(directory=resource_path("static")), name="static")

def get_db():
    db = SessionLocal()
    try:
        return db
    except:
        db.close()
        raise

def format_indian_currency(amount):
    """Formats a number in Indian currency style (e.g., 1,00,000.00)"""
    if amount is None or amount == "":
        return "0.00"
    
    try:
        amount = round(float(amount), 2)
    except (ValueError, TypeError):
        return "0.00"
        
    s = f"{amount:.2f}"
    parts = s.split('.')
    main = parts[0]
    decimal = parts[1]
    
    is_negative = main.startswith('-')
    if is_negative:
        main = main[1:]
        
    if len(main) <= 3:
        res = main
    else:
        last_three = main[-3:]
        remaining = main[:-3]
        groups = []
        while len(remaining) > 2:
            groups.append(remaining[-2:])
            remaining = remaining[:-2]
        if remaining:
            groups.append(remaining)
        res = ",".join(reversed(groups)) + "," + last_three
    
    final = f"₹{res}.{decimal}"
    return f"-{final}" if is_negative else final

# Register the filter in Jinja environment
env.filters['indian_currency'] = format_indian_currency
env.globals['now'] = datetime.now

def verify_password(plain_password, stored_password):
    return plain_password == stored_password

def is_authenticated(request: Request):
    return request.session.get("authenticated", False)

def require_auth(request: Request):
    if not is_authenticated(request):
        return RedirectResponse(url="/login", status_code=303)
    return None

@app.get("/api/names/receipts")
def get_receipt_names():
    db = SessionLocal()
    try:
        people = db.query(Person).filter(Person.person_type.in_(['Donor', 'Trust Member'])).all()
        return [{"name": p.name, "phone": p.phone or "", "address": p.address or ""} for p in people]
    finally:
        db.close()

@app.get("/api/names/vouchers")
def get_voucher_names():
    db = SessionLocal()
    try:
        people = db.query(Person).filter(Person.person_type.in_(['Beneficiary', 'Staff', 'Trust Member'])).all()
        return [{"name": p.name, "phone": p.phone or "", "address": p.address or ""} for p in people]
    finally:
        db.close()

@app.get("/api/names")
def get_unique_names():
    db = SessionLocal()
    try:
        people = db.query(Person).all()
        return [{"name": p.name, "phone": p.phone or "", "address": p.address or ""} for p in people]
    finally:
        db.close()

@app.get("/api/mad-balance/{mad_id}")
def get_mad_balance(mad_id: int):
    db = SessionLocal()
    try:
        mad = db.query(MadCategory).filter(MadCategory.id == mad_id).first()
        if not mad:
            return {"balance": 0.0, "formatted_balance": env.filters['indian_currency'](0.0)}
        
        # Calculate current balance
        receipts_total = db.query(func.sum(Receipt.amount)).filter(Receipt.mad_category_id == mad_id).scalar() or 0
        expenses_total = db.query(func.sum(Expense.amount)).filter(Expense.mad_category_id == mad_id).scalar() or 0
        
        current_balance = mad.initial_balance + receipts_total - expenses_total
        return {
            "balance": current_balance,
            "formatted_balance": env.filters['indian_currency'](current_balance)
        }
    finally:
        db.close()

@app.get("/api/payment-balance/{payment_id}")
def get_payment_balance(payment_id: int):
    db = SessionLocal()
    try:
        payment = db.query(PaymentMode).filter(PaymentMode.id == payment_id).first()
        if not payment:
            return {"balance": 0.0, "formatted_balance": env.filters['indian_currency'](0.0)}
        
        # Calculate current balance
        receipts_total = db.query(func.sum(Receipt.amount)).filter(Receipt.payment_mode_id == payment_id).scalar() or 0
        expenses_total = db.query(func.sum(Expense.amount)).filter(Expense.payment_mode_id == payment_id).scalar() or 0
        contra_in = db.query(func.sum(ContraEntry.amount)).filter(ContraEntry.to_payment_mode_id == payment_id).scalar() or 0
        contra_out = db.query(func.sum(ContraEntry.amount)).filter(ContraEntry.from_payment_mode_id == payment_id).scalar() or 0
        
        current_balance = payment.initial_balance + receipts_total - expenses_total + contra_in - contra_out
        return {
            "balance": current_balance,
            "formatted_balance": env.filters['indian_currency'](current_balance)
        }
    finally:
        db.close()

@app.get("/api/receipts")
def api_receipts(request: Request, fy: int = None, search: str = None, mad_id: int = None, register_id: int = None, payment_id: int = None):
    # Check if authenticated
    if not is_authenticated(request):
        return {"error": "Unauthorized"}
    db = get_db()
    try:
        if fy is None:
            fy = get_current_financial_year()
        start_date, end_date = get_financial_year_range(fy)
        
        query = db.query(Receipt).filter(
            and_(Receipt.date >= start_date, Receipt.date <= end_date)
        )
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                (Receipt.receipt_no.like(search_term)) | 
                (Receipt.name.like(search_term))
            )
        
        if mad_id:
            query = query.filter(Receipt.mad_category_id == mad_id)
        if register_id:
            query = query.filter(Receipt.register_id == register_id)
        if payment_id:
            query = query.filter(Receipt.payment_mode_id == payment_id)
            
        receipts = query.order_by(desc(Receipt.date)).all()
        
        return {
            "receipts": [
                {
                    "id": r.id,
                    "receipt_no": r.receipt_no,
                    "date": r.date.strftime('%d-%m-%Y'),
                    "name": r.name,
                    "phone": r.phone,
                    "address": r.address or '-',
                    "amount": r.amount,
                    "formatted_amount": env.filters['indian_currency'](r.amount),
                    "category": r.mad_category.name if r.mad_category else '-',
                    "register": r.register.name if r.register else '-',
                    "payment": r.payment_mode.name if r.payment_mode else '-'
                } for r in receipts
            ]
        }
    finally:
        db.close()

@app.get("/api/expenses")
def api_expenses(request: Request, fy: int = None, search: str = None, mad_id: int = None, expense_category_id: int = None, payment_id: int = None):
    # Check if authenticated
    if not is_authenticated(request):
        return {"error": "Unauthorized"}
    db = get_db()
    try:
        if fy is None:
            fy = get_current_financial_year()
        start_date, end_date = get_financial_year_range(fy)
        
        query = db.query(Expense).filter(
            and_(Expense.date >= start_date, Expense.date <= end_date)
        )
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                (Expense.voucher_no.like(search_term)) | 
                (Expense.name.like(search_term))
            )
        
        if mad_id:
            query = query.filter(Expense.mad_category_id == mad_id)
        if expense_category_id:
            query = query.filter(Expense.expense_category_id == expense_category_id)
        if payment_id:
            query = query.filter(Expense.payment_mode_id == payment_id)
            
        expenses = query.order_by(desc(Expense.date)).all()
        
        return {
            "expenses": [
                {
                    "id": e.id,
                    "voucher_no": e.voucher_no,
                    "date": e.date.strftime('%d-%m-%Y'),
                    "name": e.name,
                    "phone": e.phone or '-',
                    "amount": e.amount,
                    "formatted_amount": env.filters['indian_currency'](e.amount),
                    "mad_category": e.mad_category.name if e.mad_category else '-',
                    "expense_category": e.expense_category.name if e.expense_category else '-',
                    "payment": e.payment_mode.name if e.payment_mode else '-'
                } for e in expenses
            ]
        }
    finally:
        db.close()

def migrate_db():
    from sqlalchemy import text
    db = SessionLocal()
    try:
        # Check and add initial_balance to mad_categories
        try:
            db.execute(text("ALTER TABLE mad_categories ADD COLUMN initial_balance FLOAT DEFAULT 0.0"))
            db.commit()
        except Exception:
            db.rollback()
            
        # Check and add initial_balance to registers
        try:
            db.execute(text("ALTER TABLE registers ADD COLUMN initial_balance FLOAT DEFAULT 0.0"))
            db.commit()
        except Exception:
            db.rollback()
            
        # Check and add initial_balance to payment_modes
        try:
            db.execute(text("ALTER TABLE payment_modes ADD COLUMN initial_balance FLOAT DEFAULT 0.0"))
            db.commit()
        except Exception:
            db.rollback()
            
        # Check and add address to people
        try:
            db.execute(text("ALTER TABLE people ADD COLUMN address VARCHAR"))
            db.commit()
        except Exception:
            db.rollback()
    finally:
        db.close()

def migrate_people():
    db = SessionLocal()
    try:
        # Migrate Donors — only create if no profile of ANY type exists with this name
        receipts = db.query(Receipt.name, Receipt.phone, Receipt.address).all()
        for r_name, r_phone, r_address in receipts:
            if r_name:
                # First check: does a Donor profile exist?
                donor_person = db.query(Person).filter(func.lower(Person.name) == r_name.lower(), Person.person_type == 'Donor').first()
                if donor_person:
                    # Update missing fields on existing donor
                    if r_address and not donor_person.address:
                        donor_person.address = r_address
                    if r_phone and not donor_person.phone:
                        donor_person.phone = r_phone
                else:
                    # Check if ANY profile exists (e.g. Trust Member)
                    any_person = db.query(Person).filter(func.lower(Person.name) == r_name.lower()).first()
                    if not any_person:
                        db.add(Person(name=r_name, phone=r_phone or "", address=r_address or "", person_type='Donor'))
                    else:
                        # Person exists under a different type (e.g. Trust Member) — just update missing fields
                        if r_address and not any_person.address:
                            any_person.address = r_address
                        if r_phone and not any_person.phone:
                            any_person.phone = r_phone

        # Migrate Beneficiaries — only create if no profile of ANY type exists with this name
        expenses = db.query(Expense.name, Expense.phone).all()
        for e_name, e_phone in expenses:
            if e_name:
                # Check if Beneficiary/Staff profile exists
                ben_person = db.query(Person).filter(func.lower(Person.name) == e_name.lower(), Person.person_type.in_(['Beneficiary', 'Staff'])).first()
                if ben_person:
                    if e_phone and not ben_person.phone:
                        ben_person.phone = e_phone
                else:
                    # Check if ANY profile exists (e.g. Trust Member or Donor)
                    any_person = db.query(Person).filter(func.lower(Person.name) == e_name.lower()).first()
                    if not any_person:
                        db.add(Person(name=e_name, phone=e_phone or "", person_type='Beneficiary'))
                    else:
                        # Person exists under a different type — just update missing phone
                        if e_phone and not any_person.phone:
                            any_person.phone = e_phone

        db.commit()
    except Exception as e:
        print(f"Error migrating people: {e}")
        db.rollback()
    finally:
        db.close()

def deduplicate_people():
    db = SessionLocal()
    try:
        from sqlalchemy import func
        # Find duplicates by lower(name), phone, person_type
        all_people = db.query(Person).all()
        seen = {}
        for p in all_people:
            key = (p.name.lower().strip(), p.phone.strip() if p.phone else "", p.person_type)
            if key in seen:
                # Keep the first one, but if this one has an address and the first one doesn't, copy it
                first_p = seen[key]
                if p.address and not first_p.address:
                    first_p.address = p.address
                db.delete(p)
            else:
                seen[key] = p
        db.commit()
    except Exception as e:
        print(f"Error deduplicating people: {e}")
        db.rollback()
    finally:
        db.close()

@app.on_event("startup")
def startup():
    # Run migrations
    migrate_db()
    # Seed data on startup
    seed_data()
    # Migrate people from existing records
    migrate_people()
    # Deduplicate in case of repeated profiles
    deduplicate_people()

@app.get("/login")
def login_page(request: Request):
    # If already logged in, redirect to dashboard
    if is_authenticated(request):
        return RedirectResponse(url="/", status_code=303)
    template = env.get_template("login.html")
    html_content = template.render(request=request)
    return HTMLResponse(content=html_content)

@app.post("/login")
def login(request: Request, username: str = Form(...), password: str = Form(...)):
    # Verify credentials (plain text comparison)
    if username == DEFAULT_ADMIN_USERNAME and password == DEFAULT_ADMIN_PASSWORD:
        request.session["authenticated"] = True
        request.session["username"] = username
        return RedirectResponse(url="/", status_code=303)
    else:
        template = env.get_template("login.html")
        html_content = template.render(request=request, error="Invalid username or password")
        return HTMLResponse(content=html_content, status_code=401)

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse(url="/login", status_code=303)

@app.get("/change-password")
def change_password_page(request: Request):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    template = env.get_template("change_password.html")
    html_content = template.render(request=request)
    return HTMLResponse(content=html_content)

@app.post("/change-password")
def change_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...)
):
    global DEFAULT_ADMIN_PASSWORD
    
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    
    # Verify current password
    if current_password != DEFAULT_ADMIN_PASSWORD:
        template = env.get_template("change_password.html")
        html_content = template.render(request=request, error="Current password is incorrect")
        return HTMLResponse(content=html_content)
    
    # Check if new passwords match
    if new_password != confirm_password:
        template = env.get_template("change_password.html")
        html_content = template.render(request=request, error="New passwords do not match")
        return HTMLResponse(content=html_content)
    
    # Update password
    DEFAULT_ADMIN_PASSWORD = new_password
    
    template = env.get_template("change_password.html")
    html_content = template.render(request=request, success="Password changed successfully")
    return HTMLResponse(content=html_content)

@app.get("/")
def dashboard(request: Request, fy: int = None):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        # Get financial year (default to current)
        if fy is None:
            fy = get_current_financial_year()
        
        # Get date range for the financial year
        start_date, end_date = get_financial_year_range(fy)
        
        # Get total receipts and amount for selected FY
        total_receipt_count = db.query(Receipt).filter(
            and_(Receipt.date >= start_date, Receipt.date <= end_date)
        ).count()
        total_amount = db.query(func.sum(Receipt.amount)).filter(
            and_(Receipt.date >= start_date, Receipt.date <= end_date)
        ).scalar() or 0
        
        # Get total expenses for selected FY
        total_expenses = db.query(func.sum(Expense.amount)).filter(
            and_(Expense.date >= start_date, Expense.date <= end_date)
        ).scalar() or 0
        
        # Get overall balances for this FY
        opening_balance, period_total, closing_balance, period_expenses = get_balances(db, start_date, end_date)
        
        # Get recent receipts (all time)
        recent_receipts = db.query(Receipt).order_by(desc(Receipt.id)).limit(5).all()
        
        # Get Mad-wise totals for selected FY
        mad_totals = db.query(
            MadCategory.name,
            func.sum(Receipt.amount).label('total')
        ).join(Receipt).filter(
            and_(Receipt.date >= start_date, Receipt.date <= end_date)
        ).group_by(MadCategory.name).all()
        
        # Get Register-wise totals for selected FY
        register_totals = db.query(
            Register.name,
            func.sum(Receipt.amount).label('total')
        ).join(Receipt).filter(
            and_(Receipt.date >= start_date, Receipt.date <= end_date)
        ).group_by(Register.name).all()
        
        years = get_available_financial_years(db)
        
        template = env.get_template("dashboard.html")
        html_content = template.render(
            request=request,
            total_receipt_count=total_receipt_count,
            total_amount=total_amount,
            total_expenses=total_expenses,
            closing_balance=closing_balance,
            recent_receipts=recent_receipts,
            mad_totals=mad_totals,
            register_totals=register_totals,
            current_fy=fy,
            available_years=years
        )
        return HTMLResponse(content=html_content)
    finally:
        db.close()

@app.get("/add-receipt")
def add_receipt_page(request: Request):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        template = env.get_template("add_receipt.html")
        html_content = template.render(
            request=request,
            mad_categories=db.query(MadCategory).all(),
            registers=db.query(Register).all(),
            payment_modes=db.query(PaymentMode).all()
        )
        return HTMLResponse(content=html_content)
    finally:
        db.close()

@app.post("/add-receipt")
def add_receipt(
    request: Request,
    receipt_no: str = Form(...),
    date: str = Form(...),
    name: str = Form(...),
    phone: str = Form(...),
    address: str = Form(""),
    amount: float = Form(...),
    mad_category_id: int = Form(...),
    register_id: int = Form(...),
    payment_mode_id: int = Form(...),
    referred_by: str = Form(""),
    notes: str = Form("")
):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    
    try:
        # Validate phone number (exactly 10 digits)
        if not phone.isdigit() or len(phone) != 10:
            raise ValueError("Phone number must be exactly 10 digits")
        
        # Parse date
        receipt_date = datetime.strptime(date, "%Y-%m-%d").date()
        
        # Check if receipt_no already exists
        existing_receipt = db.query(Receipt).filter(Receipt.receipt_no == receipt_no).first()
        if existing_receipt:
            raise ValueError(f"Receipt number '{receipt_no}' already exists. Please check the receipt number.")
        
        # Create new receipt
        receipt = Receipt(
            receipt_no=receipt_no,
            date=receipt_date,
            name=name,
            phone=phone,
            address=address,
            amount=amount,
            mad_category_id=mad_category_id,
            register_id=register_id,
            payment_mode_id=payment_mode_id,
            referred_by=referred_by,
            notes=notes
        )
        
        db.add(receipt)
        
        # Check if person exists as Donor or Trust Member, if not create
        existing_person = db.query(Person).filter(func.lower(Person.name) == name.lower(), Person.person_type.in_(['Donor', 'Trust Member'])).first()
        if not existing_person:
            new_person = Person(name=name, phone=phone, address=address, person_type='Donor')
            db.add(new_person)
            
        db.commit()
        
        return RedirectResponse(url="/", status_code=303)
        
    except Exception as e:
        db.rollback()
        template = env.get_template("add_receipt.html")
        html_content = template.render(
            request=request,
            error=f"Error adding receipt: {str(e)}",
            mad_categories=db.query(MadCategory).all(),
            registers=db.query(Register).all(),
            payment_modes=db.query(PaymentMode).all(),
            receipt_no=receipt_no,
            date=date,
            name=name,
            phone=phone,
            address=address,
            amount=amount,
            mad_category_id=mad_category_id,
            register_id=register_id,
            payment_mode_id=payment_mode_id,
            referred_by=referred_by,
            notes=notes
        )
        return HTMLResponse(content=html_content)
    finally:
        db.close()

@app.get("/edit-receipt/{receipt_id}")
def edit_receipt_page(request: Request, receipt_id: int):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
        if not receipt:
            return RedirectResponse(url="/receipts", status_code=303)
        
        template = env.get_template("edit_receipt.html")
        html_content = template.render(
            request=request,
            receipt=receipt,
            mad_categories=db.query(MadCategory).all(),
            registers=db.query(Register).all(),
            payment_modes=db.query(PaymentMode).all()
        )
        return HTMLResponse(content=html_content)
    finally:
        db.close()

@app.post("/edit-receipt/{receipt_id}")
def edit_receipt(
    request: Request,
    receipt_id: int,
    receipt_no: str = Form(...),
    date: str = Form(...),
    name: str = Form(...),
    phone: str = Form(...),
    address: str = Form(""),
    amount: float = Form(...),
    mad_category_id: int = Form(...),
    register_id: int = Form(...),
    payment_mode_id: int = Form(...),
    referred_by: str = Form(""),
    notes: str = Form("")
):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
        if not receipt:
            return RedirectResponse(url="/receipts", status_code=303)
        
        # Validate phone number (exactly 10 digits)
        if not phone.isdigit() or len(phone) != 10:
            raise ValueError("Phone number must be exactly 10 digits")
        
        # Parse date
        receipt_date = datetime.strptime(date, "%Y-%m-%d").date()
        
        # Check if receipt_no already exists and belongs to a different receipt
        existing_receipt = db.query(Receipt).filter(Receipt.receipt_no == receipt_no).first()
        if existing_receipt and existing_receipt.id != receipt_id:
            raise ValueError(f"Receipt number '{receipt_no}' already exists. Please check the receipt number.")
        
        # Update receipt
        receipt.receipt_no = receipt_no
        receipt.date = receipt_date
        receipt.name = name
        receipt.phone = phone
        receipt.address = address
        receipt.amount = amount
        receipt.mad_category_id = mad_category_id
        receipt.register_id = register_id
        receipt.payment_mode_id = payment_mode_id
        receipt.referred_by = referred_by
        receipt.notes = notes
        
        db.commit()
        
        return RedirectResponse(url="/receipts", status_code=303)
        
    except Exception as e:
        db.rollback()
        receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
        template = env.get_template("edit_receipt.html")
        html_content = template.render(
            request=request,
            receipt=receipt,
            error=f"Error updating receipt: {str(e)}",
            mad_categories=db.query(MadCategory).all(),
            registers=db.query(Register).all(),
            payment_modes=db.query(PaymentMode).all()
        )
        return HTMLResponse(content=html_content)
    finally:
        db.close()

@app.get("/delete-receipt/{receipt_id}")
def delete_receipt_page(request: Request, receipt_id: int):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
        if not receipt:
            return RedirectResponse(url="/receipts", status_code=303)
        
        template = env.get_template("delete_confirm.html")
        html_content = template.render(
            request=request,
            receipt=receipt
        )
        return HTMLResponse(content=html_content)
    finally:
        db.close()

@app.post("/delete-receipt/{receipt_id}")
def delete_receipt(request: Request, receipt_id: int, password: str = Form(...)):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    
    # Verify password before allowing deletion
    if password != DEFAULT_ADMIN_PASSWORD:
        db = get_db()
        try:
            receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
            template = env.get_template("delete_confirm.html")
            html_content = template.render(
                request=request,
                receipt=receipt,
                error="Incorrect password. Deletion cancelled."
            )
            return HTMLResponse(content=html_content)
        finally:
            db.close()
    
    db = get_db()
    try:
        receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
        if receipt:
            db.delete(receipt)
            db.commit()
        
        return RedirectResponse(url="/receipts", status_code=303)
    finally:
        db.close()

@app.get("/receipts")
def receipts_page(request: Request, fy: int = None, search: str = None, mad_id: int = None, register_id: int = None, payment_id: int = None):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        # Get financial year (default to current)
        if fy is None:
            fy = get_current_financial_year()
        
        # Get date range for the financial year
        start_date, end_date = get_financial_year_range(fy)
        
        # Build query with F.Y. filter
        query = db.query(Receipt).filter(
            and_(Receipt.date >= start_date, Receipt.date <= end_date)
        )
        
        # Add filters
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                (Receipt.receipt_no.like(search_term)) | 
                (Receipt.name.like(search_term))
            )
        
        if mad_id:
            query = query.filter(Receipt.mad_category_id == mad_id)
        if register_id:
            query = query.filter(Receipt.register_id == register_id)
        if payment_id:
            query = query.filter(Receipt.payment_mode_id == payment_id)
            
        receipts = query.order_by(desc(Receipt.date)).all()
        
        # Calculate balances
        opening_balance, period_total, closing_balance, period_expenses = get_balances(db, start_date, end_date)
        
        years = get_available_financial_years(db)
        if not years:
            years = [fy]

        # Get master data for filters
        mad_categories = db.query(MadCategory).all()
        registers = db.query(Register).all()
        payment_modes = db.query(PaymentMode).all()
        
        template = env.get_template("receipts.html")
        html_content = template.render(
            request=request,
            receipts=receipts,
            current_fy=fy,
            available_years=years,
            search=search or "",
            mad_id=mad_id,
            register_id=register_id,
            payment_id=payment_id,
            mad_categories=mad_categories,
            registers=registers,
            payment_modes=payment_modes,
            opening_balance=opening_balance,
            period_total=period_total,
            closing_balance=closing_balance
        )
        return HTMLResponse(content=html_content)
    finally:
        db.close()

@app.get("/receipts/download")
def download_receipts_excel(request: Request, filter_type: str = "current_fy", start_date: str = None, end_date: str = None, fy: int = None, search: str = None, mad_id: int = None, register_id: int = None, payment_id: int = None):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        actual_start_date, actual_end_date, title_suffix = get_date_range_for_filter(filter_type, start_date, end_date, fy)
        
        # Build query with date filter
        query = db.query(Receipt).filter(
            and_(Receipt.date >= actual_start_date, Receipt.date <= actual_end_date)
        )
        
        # Add search filter if provided
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                (Receipt.receipt_no.like(search_term)) | 
                (Receipt.name.like(search_term))
            )
        
        # Apply other filters
        if mad_id:
            query = query.filter(Receipt.mad_category_id == mad_id)
        if register_id:
            query = query.filter(Receipt.register_id == register_id)
        if payment_id:
            query = query.filter(Receipt.payment_mode_id == payment_id)
        
        receipts = query.order_by(desc(Receipt.date)).all()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "All Receipts"
        
        # Title
        ws.merge_cells('A1:K1')
        ws['A1'] = f"All Receipts Report - {title_suffix}"
        if search:
            ws['A1'] = f"All Receipts Report - {title_suffix} (Search: {search})"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Receipt No', 'Date', 'Name', 'Phone', 'Address', 'Category', 'Register', 'Payment Mode', 'Amount', 'Referred By', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, receipt in enumerate(receipts, 4):
            ws.cell(row=row_idx, column=1, value=receipt.receipt_no)
            ws.cell(row=row_idx, column=2, value=receipt.date.strftime('%d-%m-%Y'))
            ws.cell(row=row_idx, column=3, value=receipt.name)
            ws.cell(row=row_idx, column=4, value=receipt.phone)
            ws.cell(row=row_idx, column=5, value=receipt.address or '')
            ws.cell(row=row_idx, column=6, value=receipt.mad_category.name if receipt.mad_category else '')
            ws.cell(row=row_idx, column=7, value=receipt.register.name if receipt.register else '')
            ws.cell(row=row_idx, column=8, value=receipt.payment_mode.name if receipt.payment_mode else '')
            cell = ws.cell(row=row_idx, column=9, value=receipt.amount)
            cell.number_format = '#,##,##0.00'
            ws.cell(row=row_idx, column=10, value=receipt.referred_by or '')
            ws.cell(row=row_idx, column=11, value=receipt.notes or '')
        
        # Calculate balances
        opening_balance, period_total, closing_balance, period_expenses = get_balances(db, actual_start_date, actual_end_date)
        
        # Total row
        total_row = len(receipts) + 4
        ws.cell(row=total_row, column=8, value="Total").font = Font(bold=True)
        cell = ws.cell(row=total_row, column=9, value=sum(r.amount for r in receipts))
        cell.font = Font(bold=True)
        cell.number_format = '#,##,##0.00'

        # Summary rows
        summary_row = total_row + 2
        ws.cell(row=summary_row, column=8, value="Opening Balance").font = Font(bold=True)
        ws.cell(row=summary_row, column=9, value=opening_balance).number_format = '#,##,##0.00'
        
        ws.cell(row=summary_row + 1, column=8, value="Period Total").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=9, value=period_total).number_format = '#,##,##0.00'
        
        ws.cell(row=summary_row + 2, column=8, value="Closing Balance").font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=9, value=closing_balance).number_format = '#,##,##0.00'
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 25
        
        # Save to temporary file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        search_suffix = f"_search_{search}" if search else ""
        default_filename = f"all_receipts_{title_suffix.replace(' ', '_')}{search_suffix}_{timestamp}.xlsx"
        
        filepath = prompt_save_path(
            "Save All Receipts Report",
            default_filename,
            [("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filepath:
            if not filepath.endswith('.xlsx'):
                filepath += '.xlsx'
            wb.save(filepath)
            return {"success": True, "message": f"Report saved successfully"}
        return {"success": False, "message": "Save cancelled by user."}
    except Exception as e:
        return {"success": False, "message": f"Error generating report: {str(e)}"}
    finally:
        db.close()

@app.get("/expenses/download")
def download_expenses_excel(request: Request, filter_type: str = "current_fy", start_date: str = None, end_date: str = None, fy: int = None, search: str = None, mad_id: int = None, expense_category_id: int = None, payment_id: int = None):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        actual_start_date, actual_end_date, title_suffix = get_date_range_for_filter(filter_type, start_date, end_date, fy)
        
        # Build query with date filter
        query = db.query(Expense).filter(
            and_(Expense.date >= actual_start_date, Expense.date <= actual_end_date)
        )
        
        # Add search filter if provided
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                (Expense.voucher_no.like(search_term)) | 
                (Expense.name.like(search_term))
            )
        
        # Apply other filters
        if mad_id:
            query = query.filter(Expense.mad_category_id == mad_id)
        if expense_category_id:
            query = query.filter(Expense.expense_category_id == expense_category_id)
        if payment_id:
            query = query.filter(Expense.payment_mode_id == payment_id)
        
        expenses = query.order_by(desc(Expense.date)).all()
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "All Expenses"
        
        # Title
        ws.merge_cells('A1:J1')
        ws['A1'] = f"All Expense Vouchers Report - {title_suffix}"
        if search:
            ws['A1'] = f"All Expense Vouchers Report - {title_suffix} (Search: {search})"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Voucher No', 'Date', 'Paid To', 'Phone', 'Mad Category', 'Expense Category', 'Payment Mode', 'Amount', 'Referred By', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, exp in enumerate(expenses, 4):
            ws.cell(row=row_idx, column=1, value=exp.voucher_no)
            ws.cell(row=row_idx, column=2, value=exp.date.strftime('%d-%m-%Y'))
            ws.cell(row=row_idx, column=3, value=exp.name)
            ws.cell(row=row_idx, column=4, value=exp.phone or '')
            ws.cell(row=row_idx, column=5, value=exp.mad_category.name if exp.mad_category else '')
            ws.cell(row=row_idx, column=6, value=exp.expense_category.name if exp.expense_category else '')
            ws.cell(row=row_idx, column=7, value=exp.payment_mode.name if exp.payment_mode else '')
            cell = ws.cell(row=row_idx, column=8, value=exp.amount)
            cell.number_format = '#,##,##0.00'
            ws.cell(row=row_idx, column=9, value=exp.referred_by or '')
            ws.cell(row=row_idx, column=10, value=exp.notes or '')
        
        # Total row
        total_row = len(expenses) + 4
        ws.cell(row=total_row, column=7, value="Total").font = Font(bold=True)
        cell = ws.cell(row=total_row, column=8, value=sum(e.amount for e in expenses))
        cell.font = Font(bold=True)
        cell.number_format = '#,##,##0.00'
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 25
        
        # Save to temporary file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        search_suffix = f"_search_{search}" if search else ""
        default_filename = f"all_expenses_{title_suffix.replace(' ', '_')}{search_suffix}_{timestamp}.xlsx"
        
        filepath = prompt_save_path(
            "Save All Expenses Report",
            default_filename,
            [("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filepath:
            if not filepath.endswith('.xlsx'):
                filepath += '.xlsx'
            wb.save(filepath)
            return {"success": True, "message": f"Report saved successfully"}
        return {"success": False, "message": "Save cancelled by user."}
    except Exception as e:
        return {"success": False, "message": f"Error generating report: {str(e)}"}
    finally:
        db.close()

@app.get("/mad-report")
def mad_report_page(request: Request, fy: int = None):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        # Get financial year (default to current)
        if fy is None:
            fy = get_current_financial_year()
        
        # Get date range for the financial year
        start_date, end_date = get_financial_year_range(fy)
        
        # Get all Mad categories with their initial balances
        mad_categories = db.query(MadCategory).all()
        mad_report_details = []
        for cat in mad_categories:
            # Opening receipts before this period
            opening_receipts = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.mad_category_id == cat.id, Receipt.date < start_date)
            ).scalar() or 0
            opening_expenses = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.mad_category_id == cat.id, Expense.date < start_date)
            ).scalar() or 0
            cat_opening = cat.initial_balance + opening_receipts - opening_expenses
            
            # Receipts in this period
            period_receipts = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.mad_category_id == cat.id, Receipt.date >= start_date, Receipt.date <= end_date)
            ).scalar() or 0
            
            # Expenses in this period
            period_expenses = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.mad_category_id == cat.id, Expense.date >= start_date, Expense.date <= end_date)
            ).scalar() or 0
            
            # Receipt count
            count = db.query(func.count(Receipt.id)).filter(
                and_(Receipt.mad_category_id == cat.id, Receipt.date >= start_date, Receipt.date <= end_date)
            ).scalar() or 0
            
            cat_closing = cat_opening + period_receipts - period_expenses
            
            if cat_opening != 0 or period_receipts != 0 or period_expenses != 0:
                mad_report_details.append({
                    "id": cat.id,
                    "name": cat.name,
                    "opening": cat_opening,
                    "period_in": period_receipts,
                    "period_out": period_expenses,
                    "closing": cat_closing,
                    "count": count
                })
        
        # Sort by closing balance descending
        mad_report_details = sorted(mad_report_details, key=lambda x: x["closing"], reverse=True)
        
        total_opening = sum(item["opening"] for item in mad_report_details)
        total_period = sum(item["period_in"] for item in mad_report_details)
        total_closing = sum(item["closing"] for item in mad_report_details)
        
        years = get_available_financial_years(db)
        
        template = env.get_template("mad_report.html")
        html_content = template.render(
            request=request,
            mad_report=mad_report_details,
            mad_categories_full=mad_categories,
            current_fy=fy,
            available_years=years,
            opening_balance=total_opening,
            period_total=total_period,
            closing_balance=total_closing
        )
        return HTMLResponse(content=html_content)
    finally:
        db.close()

@app.get("/mad-report/download")
def download_mad_report_excel(request: Request, filter_type: str = "current_fy", start_date: str = None, end_date: str = None, fy: int = None):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        actual_start_date, actual_end_date, title_suffix = get_date_range_for_filter(filter_type, start_date, end_date, fy)
        
        # Get detailed Mad-wise data with balances
        mad_data = []
        categories = db.query(MadCategory).all()
        for cat in categories:
            opening_r = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.mad_category_id == cat.id, Receipt.date < actual_start_date)
            ).scalar() or 0
            opening_e = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.mad_category_id == cat.id, Expense.date < actual_start_date)
            ).scalar() or 0
            opening = cat.initial_balance + opening_r - opening_e
            
            period_in = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.mad_category_id == cat.id, Receipt.date >= actual_start_date, Receipt.date <= actual_end_date)
            ).scalar() or 0
            period_out = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.mad_category_id == cat.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).scalar() or 0
            
            if opening != 0 or period_in != 0 or period_out != 0:
                mad_data.append({
                    "name": cat.name,
                    "opening": opening,
                    "period_in": period_in,
                    "period_out": period_out,
                    "closing": opening + period_in - period_out
                })
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Mad-wise Report"
        
        # Define styles
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "Mad-wise Collection Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date Range
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Date Range: {title_suffix}"
        ws['A2'].font = Font(bold=True, size=11)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ["Category Name", "Opening Balance", "Receipts (In)", "Expenses (Out)", "Closing Balance"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, item in enumerate(mad_data, start=5):
            ws.cell(row=row_idx, column=1, value=item["name"]).border = border
            
            cell_o = ws.cell(row=row_idx, column=2, value=item["opening"])
            cell_o.border = border
            cell_o.number_format = '#,##,##0.00'
            
            cell_p = ws.cell(row=row_idx, column=3, value=item["period_in"])
            cell_p.border = border
            cell_p.number_format = '#,##,##0.00'
            
            cell_out = ws.cell(row=row_idx, column=4, value=item["period_out"])
            cell_out.border = border
            cell_out.number_format = '#,##,##0.00'
            
            cell_c = ws.cell(row=row_idx, column=5, value=item["closing"])
            cell_c.border = border
            cell_c.number_format = '#,##,##0.00'
        
        # Calculate balances
        opening_balance, period_total, closing_balance, period_expenses = get_balances(db, actual_start_date, actual_end_date)
        
        # Total row
        total_row = len(mad_data) + 5
        ws.cell(row=total_row, column=1, value="Grand Total").font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = border
        
        cell_opening = ws.cell(row=total_row, column=2, value=sum(m["opening"] for m in mad_data))
        cell_opening.font = Font(bold=True)
        cell_opening.border = border
        cell_opening.number_format = '#,##,##0.00'
        
        cell_period = ws.cell(row=total_row, column=3, value=sum(m["period_in"] for m in mad_data))
        cell_period.font = Font(bold=True)
        cell_period.border = border
        cell_period.number_format = '#,##,##0.00'
        
        cell_out = ws.cell(row=total_row, column=4, value=sum(m["period_out"] for m in mad_data))
        cell_out.font = Font(bold=True)
        cell_out.border = border
        cell_out.number_format = '#,##,##0.00'
        
        cell_closing = ws.cell(row=total_row, column=5, value=sum(m["closing"] for m in mad_data))
        cell_closing.font = Font(bold=True)
        cell_closing.border = border
        cell_closing.number_format = '#,##,##0.00'

        # Summary rows
        summary_row = total_row + 2
        ws.cell(row=summary_row, column=1, value="Opening Balance").font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=opening_balance).number_format = '#,##,##0.00'
        
        ws.cell(row=summary_row + 1, column=1, value="Total Receipts").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=5, value=period_total).number_format = '#,##,##0.00'
        
        ws.cell(row=summary_row + 2, column=1, value="Total Expenses").font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=5, value=period_expenses).number_format = '#,##,##0.00'
        
        ws.cell(row=summary_row + 3, column=1, value="Closing Balance").font = Font(bold=True)
        ws.cell(row=summary_row + 3, column=5, value=closing_balance).number_format = '#,##,##0.00'
        
        # Save to file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"mad_report_{title_suffix.replace(' ', '_')}_{timestamp}.xlsx"
        
        filepath = prompt_save_path(
            "Save Mad-wise Report",
            default_filename,
            [("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filepath:
            if not filepath.endswith('.xlsx'):
                filepath += '.xlsx'
            wb.save(filepath)
            return {"success": True, "message": f"Report saved successfully"}
        return {"success": False, "message": "Save cancelled by user."}
    except Exception as e:
        return {"success": False, "message": f"Error generating report: {str(e)}"}
    finally:
        db.close()

@app.get("/expense-category-report")
def expense_category_report_page(request: Request, fy: int = None):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        # Get financial year (default to current)
        if fy is None:
            fy = get_current_financial_year()
        
        # Get date range for the financial year
        start_date, end_date = get_financial_year_range(fy)
        
        # Get all Expense categories
        expense_categories = db.query(ExpenseCategory).all()
        report_details = []
        total_period_expenses = 0
        
        for cat in expense_categories:
            # Expenses in this period
            period_out = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.expense_category_id == cat.id, Expense.date >= start_date, Expense.date <= end_date)
            ).scalar() or 0
            
            # Voucher count
            count = db.query(func.count(Expense.id)).filter(
                and_(Expense.expense_category_id == cat.id, Expense.date >= start_date, Expense.date <= end_date)
            ).scalar() or 0
            
            if period_out != 0:
                report_details.append({
                    "id": cat.id,
                    "name": cat.name,
                    "amount": period_out,
                    "count": count
                })
                total_period_expenses += period_out
        
        # Sort by amount descending
        report_details.sort(key=lambda x: x["amount"], reverse=True)
        
        years = get_available_financial_years(db)
        
        template = env.get_template("expense_category_report.html")
        return HTMLResponse(content=template.render(
            request=request,
            expense_report=report_details,
            total_expenses=total_period_expenses,
            total_vouchers=sum(item["count"] for item in report_details),
            available_years=years,
            current_fy=fy
        ))
    finally:
        db.close()

@app.get("/expense-category-report/download")
def download_expense_category_report_excel(request: Request, filter_type: str = "current_fy", start_date: str = None, end_date: str = None, fy: int = None):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        actual_start_date, actual_end_date, title_suffix = get_date_range_for_filter(filter_type, start_date, end_date, fy)
        
        # Get Expense category data
        expense_data = []
        total_sum = 0
        categories = db.query(ExpenseCategory).all()
        for cat in categories:
            amount = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.expense_category_id == cat.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).scalar() or 0
            
            count = db.query(func.count(Expense.id)).filter(
                and_(Expense.expense_category_id == cat.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).scalar() or 0
            
            if amount != 0:
                expense_data.append({
                    "name": cat.name,
                    "amount": amount,
                    "count": count
                })
                total_sum += amount
        
        # Sort by amount descending
        expense_data.sort(key=lambda x: x["amount"], reverse=True)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Expense Category Report"
        
        # Define styles
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        # Title
        ws.merge_cells('A1:C1')
        ws['A1'] = "Expense Category-wise Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date Range
        ws.merge_cells('A2:C2')
        ws['A2'] = f"Date Range: {title_suffix}"
        ws['A2'].font = Font(bold=True, size=11)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ["Category Name", "Total Expenses", "Voucher Count"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, item in enumerate(expense_data, start=5):
            ws.cell(row=row_idx, column=1, value=item["name"]).border = border
            
            cell_amt = ws.cell(row=row_idx, column=2, value=item["amount"])
            cell_amt.border = border
            cell_amt.number_format = '#,##,##0.00'
            
            ws.cell(row=row_idx, column=3, value=item["count"]).border = border
            ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal='center')
        
        # Total row
        total_row = len(expense_data) + 5
        ws.cell(row=total_row, column=1, value="Grand Total").font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = border
        
        cell_total = ws.cell(row=total_row, column=2, value=total_sum)
        cell_total.font = Font(bold=True)
        cell_total.border = border
        cell_total.number_format = '#,##,##0.00'
        
        cell_count = ws.cell(row=total_row, column=3, value=sum(i["count"] for i in expense_data))
        cell_count.font = Font(bold=True)
        cell_count.border = border
        cell_count.alignment = Alignment(horizontal='center')
        
        # Save to file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"expense_category_report_{title_suffix.replace(' ', '_')}_{timestamp}.xlsx"
        
        filepath = prompt_save_path(
            "Save Expense Category Report",
            default_filename,
            [("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filepath:
            if not filepath.endswith('.xlsx'):
                filepath += '.xlsx'
            wb.save(filepath)
            return {"success": True, "message": f"Report saved successfully"}
        return {"success": False, "message": "Save cancelled by user."}
    except Exception as e:
        return {"success": False, "message": f"Error generating report: {str(e)}"}
    finally:
        db.close()

@app.get("/register-report")
def register_report_page(request: Request, fy: int = None):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        # Get financial year (default to current)
        if fy is None:
            fy = get_current_financial_year()
        
        # Get date range for the financial year
        start_date, end_date = get_financial_year_range(fy)
        
        # Get all Registers with initial balances
        registers = db.query(Register).all()
        register_report_details = []
        for reg in registers:
            # Opening receipts before this period
            opening_receipts = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.register_id == reg.id, Receipt.date < start_date)
            ).scalar() or 0
            reg_opening = opening_receipts
            
            # Receipts in this period
            period_receipts = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.register_id == reg.id, Receipt.date >= start_date, Receipt.date <= end_date)
            ).scalar() or 0
            
            # Receipt count
            count = db.query(func.count(Receipt.id)).filter(
                and_(Receipt.register_id == reg.id, Receipt.date >= start_date, Receipt.date <= end_date)
            ).scalar() or 0
            
            reg_closing = reg_opening + period_receipts
            
            if reg_opening != 0 or period_receipts != 0:
                register_report_details.append({
                    "id": reg.id,
                    "name": reg.name,
                    "opening": reg_opening,
                    "period": period_receipts,
                    "closing": reg_closing,
                    "count": count
                })
        
        # Sort by closing balance descending
        register_report_details = sorted(register_report_details, key=lambda x: x["closing"], reverse=True)
        
        total_opening = db.query(func.sum(Receipt.amount)).filter(Receipt.date < start_date).scalar() or 0
        total_period = db.query(func.sum(Receipt.amount)).filter(
            and_(Receipt.date >= start_date, Receipt.date <= end_date)
        ).scalar() or 0
        total_closing = total_opening + total_period
        
        years = get_available_financial_years(db)
        
        template = env.get_template("register_report.html")
        html_content = template.render(
            request=request,
            register_report=register_report_details,
            registers_full=registers,
            current_fy=fy,
            available_years=years,
            opening_balance=total_opening,
            period_total=total_period,
            closing_balance=total_closing
        )
        return HTMLResponse(content=html_content)
    finally:
        db.close()

@app.get("/register-report/download")
def download_register_report_excel(request: Request, filter_type: str = "current_fy", start_date: str = None, end_date: str = None, fy: int = None):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        actual_start_date, actual_end_date, title_suffix = get_date_range_for_filter(filter_type, start_date, end_date, fy)
        
        # Get detailed Register-wise data with balances
        register_data = []
        registers = db.query(Register).all()
        for reg in registers:
            opening = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.register_id == reg.id, Receipt.date < actual_start_date)
            ).scalar() or 0
            period = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.register_id == reg.id, Receipt.date >= actual_start_date, Receipt.date <= actual_end_date)
            ).scalar() or 0
            if opening > 0 or period > 0:
                register_data.append({
                    "name": reg.name,
                    "opening": opening,
                    "period": period,
                    "closing": opening + period
                })

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Register-wise Report"
        
        # Define styles
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "Register-wise Collection Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date Range
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Date Range: {title_suffix}"
        ws['A2'].font = Font(bold=True, size=11)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ["Register Name", "Opening Balance", "Period Total", "Closing Balance"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, item in enumerate(register_data, start=5):
            ws.cell(row=row_idx, column=1, value=item["name"]).border = border
            
            cell_o = ws.cell(row=row_idx, column=2, value=item["opening"])
            cell_o.border = border
            cell_o.number_format = '#,##,##0.00'
            
            cell_p = ws.cell(row=row_idx, column=3, value=item["period"])
            cell_p.border = border
            cell_p.number_format = '#,##,##0.00'
            
            cell_c = ws.cell(row=row_idx, column=4, value=item["closing"])
            cell_c.border = border
            cell_c.number_format = '#,##,##0.00'
        
        # Total row
        total_row = len(register_data) + 5
        ws.cell(row=total_row, column=1, value="Grand Total").font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = border
        
        cell_opening = ws.cell(row=total_row, column=2, value=sum(r["opening"] for r in register_data))
        cell_opening.font = Font(bold=True)
        cell_opening.border = border
        cell_opening.number_format = '#,##,##0.00'
        
        cell_period = ws.cell(row=total_row, column=3, value=sum(r["period"] for r in register_data))
        cell_period.font = Font(bold=True)
        cell_period.border = border
        cell_period.number_format = '#,##,##0.00'
        
        cell_closing = ws.cell(row=total_row, column=4, value=sum(r["closing"] for r in register_data))
        cell_closing.font = Font(bold=True)
        cell_closing.border = border
        cell_closing.number_format = '#,##,##0.00'
        
        # Save to file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"register_report_{title_suffix.replace(' ', '_')}_{timestamp}.xlsx"
        
        filepath = prompt_save_path(
            "Save Register-wise Report",
            default_filename,
            [("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filepath:
            if not filepath.endswith('.xlsx'):
                filepath += '.xlsx'
            wb.save(filepath)
            return {"success": True, "message": f"Report saved successfully"}
        return {"success": False, "message": "Save cancelled by user."}
    except Exception as e:
        return {"success": False, "message": f"Error generating report: {str(e)}"}
    finally:
        db.close()

@app.post("/backup")
def backup_database(request: Request):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    try:
        # Create backup filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"baitulmal_backup_{timestamp}.db"
        
        filepath = prompt_save_path(
            "Save Backup As",
            backup_filename,
            [("Database files", "*.db"), ("All files", "*.*")]
        )
        
        if not filepath:
            return {"success": False, "message": "Backup cancelled by user."}
            
        if not filepath.endswith('.db'):
            filepath += '.db'
            
        # Copy database file
        shutil.copy2("baitulmal.db", filepath)
        
        # Upload to Google Drive using OAuth2
        gdrive_status = ""
        try:
            from google_auth_oauthlib.flow import InstalledAppFlow
            from google.auth.transport.requests import Request as GoogleRequest
            import pickle
            
            SCOPES = ['https://www.googleapis.com/auth/drive.file']
            
            # Check for OAuth2 credentials file (client_secret.json from Google Cloud Console)
            CLIENT_SECRET_FILE = resource_path('client_secret.json')
            TOKEN_FILE = 'token.pickle'
            
            creds = None
            
            # Load existing token
            if os.path.exists(TOKEN_FILE):
                with open(TOKEN_FILE, 'rb') as token:
                    creds = pickle.load(token)
            
            # If no valid credentials, do OAuth2 flow
            if not creds or not creds.valid:
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(GoogleRequest())
                elif os.path.exists(CLIENT_SECRET_FILE):
                    flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)
                    creds = flow.run_local_server(port=0)
                    # Save token for future runs
                    with open(TOKEN_FILE, 'wb') as token:
                        pickle.dump(creds, token)
                else:
                    gdrive_status = " (Google Drive: client_secret.json not found - run setup first)"
                    return {"success": True, "message": f"Backup saved locally. {gdrive_status}"}
            
            # Build Drive service
            service = build('drive', 'v3', credentials=creds)
            
            # Upload file to specific folder
            BACKUP_FOLDER_ID = '1okT-RuqIumUji0P94_csx_WXmj0na_q0'
            file_metadata = {
                'name': os.path.basename(filepath),
                'parents': [BACKUP_FOLDER_ID]
            }
            media = MediaFileUpload(filepath, resumable=True)
            file = service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()
            
            gdrive_status = " + uploaded to your Google Drive (Baitulmal Backups folder)"
            
        except Exception as gdrive_error:
            error_msg = str(gdrive_error)
            if "access_denied" in error_msg:
                gdrive_status = " (Google Drive: access denied - check OAuth credentials)"
            else:
                gdrive_status = f" (Google Drive: {str(gdrive_error)[:50]}...)"
        
        return {"success": True, "message": f"Backup saved locally. {gdrive_status}"}
    
    except Exception as e:
        return {"success": False, "message": f"Backup failed: {str(e)}"}

@app.post("/restore")
def restore_database(request: Request):
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    
    try:
        filepath = prompt_open_path(
            "Select Database Backup to Restore",
            [("Database files", "*.db"), ("All files", "*.*")]
        )
        
        if not filepath:
            return {"success": False, "message": "Restore cancelled by user."}
            
        # Copy selected file to baitulmal.db
        shutil.copy2(filepath, "baitulmal.db")
        
        return {"success": True, "message": f"Successfully restored from {os.path.basename(filepath)}"}
    except Exception as e:
        return {"success": False, "message": f"Restore failed: {str(e)}"}

@app.get("/payment-report")
def payment_report_page(request: Request, fy: int = None):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        # Get financial year (default to current)
        if fy is None:
            fy = get_current_financial_year()
        
        # Get date range for the financial year
        start_date, end_date = get_financial_year_range(fy)
        
        # Get Payment Mode-wise data including Contra breakdown
        payment_modes = db.query(PaymentMode).all()
        payment_report_details = []
        for mode in payment_modes:
            # Receipts in period
            receipt_total = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.payment_mode_id == mode.id, Receipt.date >= start_date, Receipt.date <= end_date)
            ).scalar() or 0
            
            # Contra In in period
            contra_in = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.to_payment_mode_id == mode.id, ContraEntry.date >= start_date, ContraEntry.date <= end_date)
            ).scalar() or 0
            
            # Contra Out in period
            contra_out = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.from_payment_mode_id == mode.id, ContraEntry.date >= start_date, ContraEntry.date <= end_date)
            ).scalar() or 0
            
            # Expenses in period
            expenses = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.payment_mode_id == mode.id, Expense.date >= start_date, Expense.date <= end_date)
            ).scalar() or 0
            
            # Opening balance for this mode
            opening_receipts = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.payment_mode_id == mode.id, Receipt.date < start_date)
            ).scalar() or 0
            opening_expenses = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.payment_mode_id == mode.id, Expense.date < start_date)
            ).scalar() or 0
            opening_contra_in = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.to_payment_mode_id == mode.id, ContraEntry.date < start_date)
            ).scalar() or 0
            opening_contra_out = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.from_payment_mode_id == mode.id, ContraEntry.date < start_date)
            ).scalar() or 0
            mode_opening = mode.initial_balance + opening_receipts - opening_expenses + opening_contra_in - opening_contra_out
            
            mode_closing = mode_opening + receipt_total - expenses + contra_in - contra_out
            
            if receipt_total != 0 or expenses != 0 or contra_in != 0 or contra_out != 0 or mode_opening != 0:
                payment_report_details.append({
                    "id": mode.id,
                    "name": mode.name,
                    "opening": mode_opening,
                    "receipts": receipt_total,
                    "expenses": expenses,
                    "contra_in": contra_in,
                    "contra_out": contra_out,
                    "closing": mode_closing
                })
        
        # Calculate balances
        opening_balance, period_total, closing_balance, period_expenses = get_balances(db, start_date, end_date)
        
        years = get_available_financial_years(db)
        
        template = env.get_template("payment_report.html")
        html_content = template.render(
            request=request,
            payment_report=payment_report_details,
            payment_modes_full=payment_modes,
            current_fy=fy,
            available_years=years,
            opening_balance=opening_balance,
            period_total=period_total,
            period_expenses=period_expenses,
            closing_balance=closing_balance
        )
        return HTMLResponse(content=html_content)
    finally:
        db.close()

@app.get("/payment-report/download")
def download_payment_report_excel(request: Request, filter_type: str = "current_fy", start_date: str = None, end_date: str = None, fy: int = None):
    # Check if authenticated
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        actual_start_date, actual_end_date, title_suffix = get_date_range_for_filter(filter_type, start_date, end_date, fy)
        
        # Get detailed Payment Mode-wise data with balances
        payment_data = []
        modes = db.query(PaymentMode).all()
        for mode in modes:
            opening_r = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.payment_mode_id == mode.id, Receipt.date < actual_start_date)
            ).scalar() or 0
            opening_e = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.payment_mode_id == mode.id, Expense.date < actual_start_date)
            ).scalar() or 0
            opening_cin = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.to_payment_mode_id == mode.id, ContraEntry.date < actual_start_date)
            ).scalar() or 0
            opening_cout = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.from_payment_mode_id == mode.id, ContraEntry.date < actual_start_date)
            ).scalar() or 0
            opening = mode.initial_balance + opening_r - opening_e + opening_cin - opening_cout
            
            period_in = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.payment_mode_id == mode.id, Receipt.date >= actual_start_date, Receipt.date <= actual_end_date)
            ).scalar() or 0
            period_out = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.payment_mode_id == mode.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).scalar() or 0
            contra_in = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.to_payment_mode_id == mode.id, ContraEntry.date >= actual_start_date, ContraEntry.date <= actual_end_date)
            ).scalar() or 0
            contra_out = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.from_payment_mode_id == mode.id, ContraEntry.date >= actual_start_date, ContraEntry.date <= actual_end_date)
            ).scalar() or 0
            
            if opening != 0 or period_in != 0 or period_out != 0 or contra_in != 0 or contra_out != 0:
                payment_data.append({
                    "name": mode.name,
                    "opening": opening,
                    "receipts": period_in,
                    "expenses": period_out,
                    "contra_in": contra_in,
                    "contra_out": contra_out,
                    "closing": opening + period_in - period_out + contra_in - contra_out
                })

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payment Report"
        
        # Define styles
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = "Payment Mode-wise Collection & Expense Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date Range
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Date Range: {title_suffix}"
        ws['A2'].font = Font(bold=True, size=11)
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ["Payment Mode", "Opening Balance", "Receipts", "Expenses", "Contra In", "Contra Out", "Closing Balance"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, item in enumerate(payment_data, start=5):
            ws.cell(row=row_idx, column=1, value=item["name"]).border = border
            
            cell_o = ws.cell(row=row_idx, column=2, value=item["opening"])
            cell_o.border = border
            cell_o.number_format = '#,##,##0.00'
            
            cell_r = ws.cell(row=row_idx, column=3, value=item["receipts"])
            cell_r.border = border
            cell_r.number_format = '#,##,##0.00'
            
            cell_e = ws.cell(row=row_idx, column=4, value=item["expenses"])
            cell_e.border = border
            cell_e.number_format = '#,##,##0.00'
            
            cell_cin = ws.cell(row=row_idx, column=5, value=item["contra_in"])
            cell_cin.border = border
            cell_cin.number_format = '#,##,##0.00'
            
            cell_cout = ws.cell(row=row_idx, column=6, value=item["contra_out"])
            cell_cout.border = border
            cell_cout.number_format = '#,##,##0.00'
            
            cell_c = ws.cell(row=row_idx, column=7, value=item["closing"])
            cell_c.border = border
            cell_c.number_format = '#,##,##0.00'
        
        # Total row
        total_row = len(payment_data) + 5
        ws.cell(row=total_row, column=1, value="Grand Total").font = Font(bold=True)
        ws.cell(row=total_row, column=1).border = border
        
        cell_opening = ws.cell(row=total_row, column=2, value=sum(p["opening"] for p in payment_data))
        cell_opening.font = Font(bold=True)
        cell_opening.border = border
        cell_opening.number_format = '#,##,##0.00'
        
        cell_r = ws.cell(row=total_row, column=3, value=sum(p["receipts"] for p in payment_data))
        cell_r.font = Font(bold=True)
        cell_r.border = border
        cell_r.number_format = '#,##,##0.00'
        
        cell_e = ws.cell(row=total_row, column=4, value=sum(p["expenses"] for p in payment_data))
        cell_e.font = Font(bold=True)
        cell_e.border = border
        cell_e.number_format = '#,##,##0.00'
        
        cell_cin = ws.cell(row=total_row, column=5, value=sum(p["contra_in"] for p in payment_data))
        cell_cin.font = Font(bold=True)
        cell_cin.border = border
        cell_cin.number_format = '#,##,##0.00'
        
        cell_cout = ws.cell(row=total_row, column=6, value=sum(p["contra_out"] for p in payment_data))
        cell_cout.font = Font(bold=True)
        cell_cout.border = border
        cell_cout.number_format = '#,##,##0.00'
        
        cell_closing = ws.cell(row=total_row, column=7, value=sum(p["closing"] for p in payment_data))
        cell_closing.font = Font(bold=True)
        cell_closing.border = border
        cell_closing.number_format = '#,##,##0.00'
        
        
        # Save to file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"payment_report_{title_suffix.replace(' ', '_')}_{timestamp}.xlsx"
        
        filepath = prompt_save_path(
            "Save Payment Report",
            default_filename,
            [("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filepath:
            if not filepath.endswith('.xlsx'):
                filepath += '.xlsx'
            wb.save(filepath)
            return {"success": True, "message": f"Report saved successfully"}
        return {"success": False, "message": "Save cancelled by user."}
    except Exception as e:
        return {"success": False, "message": f"Error generating report: {str(e)}"}
    finally:
        db.close()

@app.get("/audit-report")
def audit_report_page(request: Request, fy: str = None, start_date: str = None, end_date: str = None, filter_type: str = "current_fy"):
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    
    db = get_db()
    try:
        actual_start_date, actual_end_date, title_suffix = get_date_range_for_filter(filter_type, start_date, end_date, fy)
        
        # 1. Calculate Balances for Payment Modes
        payment_modes = db.query(PaymentMode).all()
        payment_data = []
        
        for mode in payment_modes:
            opening_r = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.payment_mode_id == mode.id, Receipt.date < actual_start_date)
            ).scalar() or 0
            opening_e = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.payment_mode_id == mode.id, Expense.date < actual_start_date)
            ).scalar() or 0
            opening_cin = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.to_payment_mode_id == mode.id, ContraEntry.date < actual_start_date)
            ).scalar() or 0
            opening_cout = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.from_payment_mode_id == mode.id, ContraEntry.date < actual_start_date)
            ).scalar() or 0
            opening = mode.initial_balance + opening_r - opening_e + opening_cin - opening_cout
            
            period_in = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.payment_mode_id == mode.id, Receipt.date >= actual_start_date, Receipt.date <= actual_end_date)
            ).scalar() or 0
            period_out = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.payment_mode_id == mode.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).scalar() or 0
            contra_in = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.to_payment_mode_id == mode.id, ContraEntry.date >= actual_start_date, ContraEntry.date <= actual_end_date)
            ).scalar() or 0
            contra_out = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.from_payment_mode_id == mode.id, ContraEntry.date >= actual_start_date, ContraEntry.date <= actual_end_date)
            ).scalar() or 0
            
            closing = opening + period_in - period_out + contra_in - contra_out
            
            payment_data.append({
                "id": mode.id,
                "name": mode.name,
                "opening": opening,
                "receipts": period_in,
                "expenses": period_out,
                "contra_in": contra_in,
                "contra_out": contra_out,
                "closing": closing
            })
            
        # 2. Calculate Balances for MAD Categories
        categories = db.query(MadCategory).all()
        mad_data = []
        for cat in categories:
            opening_r = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.mad_category_id == cat.id, Receipt.date < actual_start_date)
            ).scalar() or 0
            opening_e = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.mad_category_id == cat.id, Expense.date < actual_start_date)
            ).scalar() or 0
            opening = cat.initial_balance + opening_r - opening_e
            
            period_in = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.mad_category_id == cat.id, Receipt.date >= actual_start_date, Receipt.date <= actual_end_date)
            ).scalar() or 0
            period_out = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.mad_category_id == cat.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).scalar() or 0
            closing = opening + period_in - period_out
            
            mad_data.append({
                "id": cat.id,
                "name": cat.name,
                "opening": opening,
                "period_in": period_in,
                "period_out": period_out,
                "closing": closing
            })
            
        # 3. Calculate Expense Categories Breakdown
        expense_cats = db.query(ExpenseCategory).all()
        expense_data = []
        total_revenue_expenses = 0
        for ec in expense_cats:
            spent = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.expense_category_id == ec.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).scalar() or 0
            count = db.query(func.count(Expense.id)).filter(
                and_(Expense.expense_category_id == ec.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).scalar() or 0
            
            if spent > 0 or count > 0:
                expense_data.append({
                    "id": ec.id,
                    "name": ec.name,
                    "amount": spent,
                    "count": count,
                    "classification": classify_expense(ec.name)
                })
                total_revenue_expenses += spent
                
        expense_data.sort(key=lambda x: x["amount"], reverse=True)
        
        # 4. Fixed Assets Details (Capital Outlay)
        fixed_assets_cat = db.query(ExpenseCategory).filter(ExpenseCategory.name.ilike('%fixed asset%')).first()
        fixed_assets_list = []
        total_fixed_assets_amount = 0
        if fixed_assets_cat:
            assets = db.query(Expense).filter(
                and_(Expense.expense_category_id == fixed_assets_cat.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).order_by(Expense.date).all()
            for asset in assets:
                fixed_assets_list.append({
                    "id": asset.id,
                    "date": asset.date.strftime('%Y-%m-%d'),
                    "voucher_no": asset.voucher_no,
                    "name": asset.name,
                    "description": asset.notes or "Fixed asset acquisition",
                    "amount": asset.amount
                })
                total_fixed_assets_amount += asset.amount

        # 5. Calculate Compliance Indicators
        # A. Administrative Overhead Ratio (using central classification map)
        total_admin_expenses = sum(item["amount"] for item in expense_data if item["classification"] == "admin")
        total_charity_expenses = sum(item["amount"] for item in expense_data if item["classification"] == "charity")
        total_asset_expenses = sum(item["amount"] for item in expense_data if item["classification"] == "asset")
        admin_ratio = (total_admin_expenses / total_revenue_expenses * 100) if total_revenue_expenses > 0 else 0
        
        # B. Cash vs Bank Ratio
        total_closing_payment_modes = sum(mode["closing"] for mode in payment_data)
        total_cash_closing = sum(mode["closing"] for mode in payment_data if "cash" in mode["name"].lower())
        total_bank_closing = total_closing_payment_modes - total_cash_closing
        cash_ratio = (total_cash_closing / total_closing_payment_modes * 100) if total_closing_payment_modes > 0 else 0
        bank_ratio = (total_bank_closing / total_closing_payment_modes * 100) if total_closing_payment_modes > 0 else 0
        
        # C. Audit Trail Completeness (Missing Phone numbers)
        total_receipts_count = db.query(func.count(Receipt.id)).filter(
            and_(Receipt.date >= actual_start_date, Receipt.date <= actual_end_date)
        ).scalar() or 0
        receipts_missing_phone = db.query(func.count(Receipt.id)).filter(
            and_(Receipt.date >= actual_start_date, Receipt.date <= actual_end_date, (Receipt.phone == '') | (Receipt.phone == None))
        ).scalar() or 0
        
        total_expenses_count = db.query(func.count(Expense.id)).filter(
            and_(Expense.date >= actual_start_date, Expense.date <= actual_end_date)
        ).scalar() or 0
        expenses_missing_phone = db.query(func.count(Expense.id)).filter(
            and_(Expense.date >= actual_start_date, Expense.date <= actual_end_date, (Expense.phone == '') | (Expense.phone == None))
        ).scalar() or 0
        
        total_tx_count = total_receipts_count + total_expenses_count
        total_missing_phone = receipts_missing_phone + expenses_missing_phone
        trail_completeness = ((total_tx_count - total_missing_phone) / total_tx_count * 100) if total_tx_count > 0 else 100.0
        
        # D. Reconciliation Check
        total_mads_closing = sum(mad["closing"] for mad in mad_data)
        reconciliation_discrepancy = total_closing_payment_modes - total_mads_closing
        reconciled = abs(reconciliation_discrepancy) < 0.01
        
        # E. Summary Totals
        total_opening = sum(mode["opening"] for mode in payment_data)
        total_receipts = sum(mode["receipts"] for mode in payment_data)
        total_expenses = sum(mode["expenses"] for mode in payment_data)
        
        years = get_available_financial_years(db)
        
        # Sanitize fy for the template
        display_fy = get_current_financial_year()
        if fy is not None:
            try:
                display_fy = int(str(fy).split('?')[0].split('&')[0])
            except ValueError:
                pass
        
        template = env.get_template("audit_report.html")
        return HTMLResponse(content=template.render(
            request=request,
            filter_type=filter_type,
            start_date=actual_start_date.strftime('%Y-%m-%d'),
            end_date=actual_end_date.strftime('%Y-%m-%d'),
            current_fy=display_fy,
            available_years=years,
            title_suffix=title_suffix,
            payment_modes=payment_data,
            mad_categories=mad_data,
            expense_categories=expense_data,
            fixed_assets=fixed_assets_list,
            total_fixed_assets_amount=total_fixed_assets_amount,
            total_opening=total_opening,
            total_receipts=total_receipts,
            total_expenses=total_expenses,
            total_closing=total_closing_payment_modes,
            total_mads_closing=total_mads_closing,
            reconciliation_discrepancy=reconciliation_discrepancy,
            reconciled=reconciled,
            total_admin_expenses=total_admin_expenses,
            total_charity_expenses=total_charity_expenses,
            total_asset_expenses=total_asset_expenses,
            admin_ratio=admin_ratio,
            total_cash_closing=total_cash_closing,
            total_bank_closing=total_bank_closing,
            cash_ratio=cash_ratio,
            bank_ratio=bank_ratio,
            total_tx_count=total_tx_count,
            total_missing_phone=total_missing_phone,
            trail_completeness=trail_completeness,
            now=datetime.now()
        ))
    finally:
        db.close()

@app.get("/audit-report/download")
def download_audit_report_excel(request: Request, filter_type: str = "current_fy", start_date: str = None, end_date: str = None, fy: str = None):
    auth_redirect = require_auth(request)
    if auth_redirect:
        return auth_redirect
    db = get_db()
    try:
        actual_start_date, actual_end_date, title_suffix = get_date_range_for_filter(filter_type, start_date, end_date, fy)
        
        # 1. Fetch Payment Mode data
        payment_modes = db.query(PaymentMode).all()
        payment_data = []
        for mode in payment_modes:
            opening_r = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.payment_mode_id == mode.id, Receipt.date < actual_start_date)
            ).scalar() or 0
            opening_e = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.payment_mode_id == mode.id, Expense.date < actual_start_date)
            ).scalar() or 0
            opening_cin = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.to_payment_mode_id == mode.id, ContraEntry.date < actual_start_date)
            ).scalar() or 0
            opening_cout = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.from_payment_mode_id == mode.id, ContraEntry.date < actual_start_date)
            ).scalar() or 0
            opening = mode.initial_balance + opening_r - opening_e + opening_cin - opening_cout
            
            period_in = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.payment_mode_id == mode.id, Receipt.date >= actual_start_date, Receipt.date <= actual_end_date)
            ).scalar() or 0
            period_out = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.payment_mode_id == mode.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).scalar() or 0
            contra_in = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.to_payment_mode_id == mode.id, ContraEntry.date >= actual_start_date, ContraEntry.date <= actual_end_date)
            ).scalar() or 0
            contra_out = db.query(func.sum(ContraEntry.amount)).filter(
                and_(ContraEntry.from_payment_mode_id == mode.id, ContraEntry.date >= actual_start_date, ContraEntry.date <= actual_end_date)
            ).scalar() or 0
            
            closing = opening + period_in - period_out + contra_in - contra_out
            payment_data.append({
                "name": mode.name,
                "opening": opening,
                "receipts": period_in,
                "expenses": period_out,
                "contra_in": contra_in,
                "contra_out": contra_out,
                "closing": closing
            })
            
        # 2. Fetch MAD data
        categories = db.query(MadCategory).all()
        mad_data = []
        for cat in categories:
            opening_r = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.mad_category_id == cat.id, Receipt.date < actual_start_date)
            ).scalar() or 0
            opening_e = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.mad_category_id == cat.id, Expense.date < actual_start_date)
            ).scalar() or 0
            opening = cat.initial_balance + opening_r - opening_e
            
            period_in = db.query(func.sum(Receipt.amount)).filter(
                and_(Receipt.mad_category_id == cat.id, Receipt.date >= actual_start_date, Receipt.date <= actual_end_date)
            ).scalar() or 0
            period_out = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.mad_category_id == cat.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).scalar() or 0
            closing = opening + period_in - period_out
            mad_data.append({
                "name": cat.name,
                "opening": opening,
                "period_in": period_in,
                "period_out": period_out,
                "closing": closing
            })
            
        # 3. Fetch Expense breakdown
        expense_cats = db.query(ExpenseCategory).all()
        expense_data = []
        total_revenue_expenses = 0
        for ec in expense_cats:
            spent = db.query(func.sum(Expense.amount)).filter(
                and_(Expense.expense_category_id == ec.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).scalar() or 0
            count = db.query(func.count(Expense.id)).filter(
                and_(Expense.expense_category_id == ec.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).scalar() or 0
            if spent > 0 or count > 0:
                expense_data.append({
                    "name": ec.name,
                    "amount": spent,
                    "count": count,
                    "classification": classify_expense(ec.name)
                })
                total_revenue_expenses += spent
        expense_data.sort(key=lambda x: x["amount"], reverse=True)
        
        # 4. Fetch Fixed Assets details
        fixed_assets_cat = db.query(ExpenseCategory).filter(ExpenseCategory.name.ilike('%fixed asset%')).first()
        fixed_assets_list = []
        total_fixed_assets_amount = 0
        if fixed_assets_cat:
            assets = db.query(Expense).filter(
                and_(Expense.expense_category_id == fixed_assets_cat.id, Expense.date >= actual_start_date, Expense.date <= actual_end_date)
            ).order_by(Expense.date).all()
            for asset in assets:
                fixed_assets_list.append({
                    "date": asset.date.strftime('%Y-%m-%d'),
                    "voucher_no": asset.voucher_no,
                    "name": asset.name,
                    "description": asset.notes or "Fixed asset acquisition",
                    "amount": asset.amount
                })
                total_fixed_assets_amount += asset.amount
                
        # 5. Compliance calculations
        admin_keywords = ["salary", "rent", "office", "admin", "electric", "telephon", "stationery", "net", "wifi", "maintenance", "bank charges", "repair"]
        total_admin_expenses = sum(item["amount"] for item in expense_data if any(kw in item["name"].lower() for kw in admin_keywords))
        admin_ratio = (total_admin_expenses / total_revenue_expenses * 100) if total_revenue_expenses > 0 else 0
        
        total_closing_payment_modes = sum(mode["closing"] for mode in payment_data)
        total_cash_closing = sum(mode["closing"] for mode in payment_data if "cash" in mode["name"].lower())
        total_bank_closing = total_closing_payment_modes - total_cash_closing
        cash_ratio = (total_cash_closing / total_closing_payment_modes * 100) if total_closing_payment_modes > 0 else 0
        bank_ratio = (total_bank_closing / total_closing_payment_modes * 100) if total_closing_payment_modes > 0 else 0
        
        total_receipts_count = db.query(func.count(Receipt.id)).filter(
            and_(Receipt.date >= actual_start_date, Receipt.date <= actual_end_date)
        ).scalar() or 0
        receipts_missing_phone = db.query(func.count(Receipt.id)).filter(
            and_(Receipt.date >= actual_start_date, Receipt.date <= actual_end_date, (Receipt.phone == '') | (Receipt.phone == None))
        ).scalar() or 0
        total_expenses_count = db.query(func.count(Expense.id)).filter(
            and_(Expense.date >= actual_start_date, Expense.date <= actual_end_date)
        ).scalar() or 0
        expenses_missing_phone = db.query(func.count(Expense.id)).filter(
            and_(Expense.date >= actual_start_date, Expense.date <= actual_end_date, (Expense.phone == '') | (Expense.phone == None))
        ).scalar() or 0
        total_tx_count = total_receipts_count + total_expenses_count
        total_missing_phone = receipts_missing_phone + expenses_missing_phone
        trail_completeness = ((total_tx_count - total_missing_phone) / total_tx_count * 100) if total_tx_count > 0 else 100.0
        
        total_mads_closing = sum(mad["closing"] for mad in mad_data)
        reconciliation_discrepancy = total_closing_payment_modes - total_mads_closing
        reconciled_status = "Reconciled" if abs(reconciliation_discrepancy) < 0.01 else f"Discrepancy of ₹{reconciliation_discrepancy:.2f}"
        
        # Generate Receipts and Payments Data
        receipts_lines = []
        payments_lines = []
        
        # Receipts - Opening Balance
        receipts_lines.append(("To OPENING BALANCE:", "", True))
        
        cash_opening = sum(m["opening"] for m in payment_data if "cash" in m["name"].lower())
        receipts_lines.append(("   Cash In Hand", cash_opening, False))
        
        bank_modes = [m for m in payment_data if "cash" not in m["name"].lower() and m["opening"] > 0]
        if bank_modes:
            receipts_lines.append(("   Cash at Bank's:", "", True))
            for m in bank_modes:
                receipts_lines.append((f"      {m['name']}", m["opening"], False))
                
        # Receipts - Collections
        receipts_lines.append(("\" COLLECTION FROM:", "", True))
        for m in mad_data:
            if m["period_in"] > 0:
                receipts_lines.append((f"   {m['name']}", m["period_in"], False))
                
        # Payments - Charity
        charity_exps = [e for e in expense_data if e["classification"] == "charity" and e["amount"] > 0]
        if charity_exps:
            payments_lines.append(("By CHARITY PROGRAMMES:", "", True))
            for e in charity_exps:
                payments_lines.append((f"   {e['name']}", e["amount"], False))
                
        # Payments - Admin
        admin_exps = [e for e in expense_data if e["classification"] == "admin" and e["amount"] > 0]
        if admin_exps:
            payments_lines.append(("\" ADMINISTRATIVE EXPENSES:", "", True))
            for e in admin_exps:
                payments_lines.append((f"   {e['name']}", e["amount"], False))
                
        # Payments - Asset
        asset_exps = [e for e in expense_data if e["classification"] == "asset" and e["amount"] > 0]
        if asset_exps:
            payments_lines.append(("\" PURCHASES:", "", True))
            for e in asset_exps:
                payments_lines.append((f"   {e['name']}", e["amount"], False))
                
        # Payments - Closing Balance
        payments_lines.append(("\" CLOSING BALANCE:", "", True))
        cash_closing = sum(m["closing"] for m in payment_data if "cash" in m["name"].lower())
        payments_lines.append(("   Cash In Hand", cash_closing, False))
        
        bank_modes_c = [m for m in payment_data if "cash" not in m["name"].lower() and m["closing"] > 0]
        if bank_modes_c:
            payments_lines.append(("   Cash at Bank's:", "", True))
            for m in bank_modes_c:
                payments_lines.append((f"      {m['name']}", m["closing"], False))
                
        # Calculate totals
        total_receipts = sum(l[1] for l in receipts_lines if isinstance(l[1], (int, float)))
        total_payments = sum(l[1] for l in payments_lines if isinstance(l[1], (int, float)))
        
        # Pad lines
        max_len = max(len(receipts_lines), len(payments_lines))
        while len(receipts_lines) < max_len:
            receipts_lines.append(("", "", False))
        while len(payments_lines) < max_len:
            payments_lines.append(("", "", False))
            
        # Workbook setup
        wb = Workbook()
        ws = wb.active
        ws.title = "Receipts and Payments"
        
        # Styles
        title_font = Font(name="Calibri", size=14, bold=True)
        subtitle_font = Font(name="Calibri", size=12)
        header_font = Font(name="Calibri", size=11, bold=True)
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)
        
        thin_border_top_bottom = Border(top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
        double_bottom_border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
        
        ws.merge_cells('A1:D1')
        ws['A1'] = "HAZRATH SHER-E-SAWAR (RH) BAITULMAL TRUST"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:D2')
        ws['A2'] = "DARGAH HAZRATH SHER-E-SAWAR (RH), BASAVAKALYAN, DISTRICT: BIDAR"
        ws['A2'].font = subtitle_font
        ws['A2'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A4:D4')
        ws['A4'] = f"RECEIPTS AND PAYMENTS ACCOUNT FOR THE PERIOD {title_suffix.upper()}"
        ws['A4'].font = header_font
        ws['A4'].alignment = Alignment(horizontal='center')
        
        # Table Headers
        headers = [("A5", "RECEIPTS"), ("B5", "AMOUNT"), ("C5", "PAYMENTS"), ("D5", "AMOUNT")]
        for cell_ref, val in headers:
            ws[cell_ref] = val
            ws[cell_ref].font = bold_font
            ws[cell_ref].border = thin_border_top_bottom
            if 'B' in cell_ref or 'D' in cell_ref:
                ws[cell_ref].alignment = Alignment(horizontal='right')
                
        # Data Rows
        current_row = 6
        for i in range(max_len):
            r_label, r_amount, r_bold = receipts_lines[i]
            p_label, p_amount, p_bold = payments_lines[i]
            
            cell_r_label = ws.cell(row=current_row, column=1, value=r_label)
            cell_r_amount = ws.cell(row=current_row, column=2, value=r_amount if r_amount != "" else None)
            cell_p_label = ws.cell(row=current_row, column=3, value=p_label)
            cell_p_amount = ws.cell(row=current_row, column=4, value=p_amount if p_amount != "" else None)
            
            cell_r_label.font = bold_font if r_bold else regular_font
            cell_p_label.font = bold_font if p_bold else regular_font
            cell_r_amount.font = regular_font
            cell_p_amount.font = regular_font
            
            if isinstance(r_amount, (int, float)):
                cell_r_amount.number_format = '#,##,##0.00'
            if isinstance(p_amount, (int, float)):
                cell_p_amount.number_format = '#,##,##0.00'
                
            current_row += 1
            
        # Totals Row
        ws.cell(row=current_row, column=1, value="")
        ws.cell(row=current_row, column=3, value="")
        
        c_tr = ws.cell(row=current_row, column=2, value=total_receipts)
        c_tr.font = bold_font
        c_tr.border = double_bottom_border
        c_tr.number_format = '#,##,##0.00'
        
        c_tp = ws.cell(row=current_row, column=4, value=total_payments)
        c_tp.font = bold_font
        c_tp.border = double_bottom_border
        c_tp.number_format = '#,##,##0.00'
        
        current_row += 2
        ws.cell(row=current_row, column=1, value="\"SUBJECT TO SEPARATE REPORT OF EVEN DATE\".").font = regular_font
        
        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 20
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"Receipts_Payments_Account_{title_suffix.replace(' ', '_')}_{timestamp}.xlsx"
        
        filepath = prompt_save_path(
            "Save Receipts and Payments Account",
            default_filename,
            [("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if filepath:
            if not filepath.endswith('.xlsx'):
                filepath += '.xlsx'
            wb.save(filepath)
            return {"success": True, "message": "Excel report saved successfully."}
        return {"success": False, "message": "Save cancelled."}
        
    except Exception as e:
        return {"success": False, "message": f"Error generating Excel report: {str(e)}"}
    finally:
        db.close()




@app.get("/contra")
def contra_page(request: Request, fy: int = None):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    db = get_db()
    try:
        if fy is None: fy = get_current_financial_year()
        start_date, end_date = get_financial_year_range(fy)
        
        entries = db.query(ContraEntry).filter(
            and_(ContraEntry.date >= start_date, ContraEntry.date <= end_date)
        ).order_by(desc(ContraEntry.date)).all()
        
        payment_modes = db.query(PaymentMode).all()
        
        template = env.get_template("contra.html")
        return HTMLResponse(content=template.render(
            request=request,
            entries=entries,
            payment_modes=payment_modes,
            current_fy=fy
        ))
    finally:
        db.close()

@app.post("/add-contra")
def add_contra(
    request: Request,
    date: str = Form(...),
    from_mode_id: int = Form(...),
    to_mode_id: int = Form(...),
    amount: float = Form(...),
    notes: str = Form("")
):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    
    db = get_db()
    try:
        entry = ContraEntry(
            date=datetime.strptime(date, "%Y-%m-%d").date(),
            from_payment_mode_id=from_mode_id,
            to_payment_mode_id=to_mode_id,
            amount=amount,
            notes=notes
        )
        db.add(entry)
        db.commit()
        return RedirectResponse(url="/contra", status_code=303)
    finally:
        db.close()

@app.get("/delete-contra/{contra_id}")
def delete_contra_confirm(request: Request, contra_id: int):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    db = get_db()
    try:
        entry = db.query(ContraEntry).filter(ContraEntry.id == contra_id).first()
        if not entry:
            return RedirectResponse(url="/contra", status_code=303)
        
        template = env.get_template("delete_contra_confirm.html")
        return HTMLResponse(content=template.render(
            request=request,
            entry=entry
        ))
    finally:
        db.close()

@app.post("/delete-contra/{contra_id}")
def delete_contra(request: Request, contra_id: int, password: str = Form(...)):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    
    # Verify password
    if password != DEFAULT_ADMIN_PASSWORD:
        db = get_db()
        try:
            entry = db.query(ContraEntry).filter(ContraEntry.id == contra_id).first()
            template = env.get_template("delete_contra_confirm.html")
            return HTMLResponse(content=template.render(
                request=request,
                entry=entry,
                error="Incorrect password. Deletion cancelled."
            ))
        finally:
            db.close()

    db = get_db()
    try:
        entry = db.query(ContraEntry).filter(ContraEntry.id == contra_id).first()
        if entry:
            db.delete(entry)
            db.commit()
        return RedirectResponse(url="/contra", status_code=303)
    finally:
        db.close()

@app.get("/edit-contra/{contra_id}")
def edit_contra_page(request: Request, contra_id: int):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    db = get_db()
    try:
        entry = db.query(ContraEntry).filter(ContraEntry.id == contra_id).first()
        if not entry:
            return RedirectResponse(url="/contra", status_code=303)
        
        payment_modes = db.query(PaymentMode).all()
        template = env.get_template("edit_contra.html")
        return HTMLResponse(content=template.render(
            request=request,
            entry=entry,
            payment_modes=payment_modes
        ))
    finally:
        db.close()

@app.post("/edit-contra/{contra_id}")
def edit_contra(
    request: Request,
    contra_id: int,
    date: str = Form(...),
    from_mode_id: int = Form(...),
    to_mode_id: int = Form(...),
    amount: float = Form(...),
    notes: str = Form("")
):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    
    db = get_db()
    try:
        entry = db.query(ContraEntry).filter(ContraEntry.id == contra_id).first()
        if entry:
            entry.date = datetime.strptime(date, "%Y-%m-%d").date()
            entry.from_payment_mode_id = from_mode_id
            entry.to_payment_mode_id = to_mode_id
            entry.amount = amount
            entry.notes = notes
            db.commit()
        return RedirectResponse(url="/contra", status_code=303)
    finally:
        db.close()

@app.get("/expenses")
def expenses_page(request: Request, fy: int = None, search: str = None, mad_id: int = None, expense_category_id: int = None, payment_id: int = None):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    db = get_db()
    try:
        if fy is None: fy = get_current_financial_year()
        start_date, end_date = get_financial_year_range(fy)
        
        query = db.query(Expense).filter(and_(Expense.date >= start_date, Expense.date <= end_date))
        
        if search:
            search_term = f"%{search}%"
            query = query.filter((Expense.name.like(search_term)) | (Expense.voucher_no.like(search_term)))
            
        if mad_id:
            query = query.filter(Expense.mad_category_id == mad_id)
        if expense_category_id:
            query = query.filter(Expense.expense_category_id == expense_category_id)
        if payment_id:
            query = query.filter(Expense.payment_mode_id == payment_id)
            
        expenses = query.order_by(desc(Expense.date)).all()
        years = get_available_financial_years(db)
        
        template = env.get_template("expenses.html")
        return HTMLResponse(content=template.render(
            request=request,
            expenses=expenses,
            current_fy=fy,
            available_years=years,
            search=search or "",
            mad_id=mad_id,
            expense_category_id=expense_category_id,
            payment_id=payment_id,
            mad_categories=db.query(MadCategory).all(),
            expense_categories=db.query(ExpenseCategory).all(),
            payment_modes=db.query(PaymentMode).all()
        ))
    finally:
        db.close()

@app.get("/add-voucher")
def add_voucher_page(request: Request):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    db = get_db()
    try:
        template = env.get_template("add_voucher.html")
        return HTMLResponse(content=template.render(
            request=request,
            mad_categories=db.query(MadCategory).all(),
            expense_categories=db.query(ExpenseCategory).all(),
            payment_modes=db.query(PaymentMode).all()
        ))
    finally:
        db.close()

@app.post("/add-voucher")
async def add_voucher(
    request: Request,
    voucher_no: str = Form(...),
    date: str = Form(...),
    name: str = Form(...),
    phone: str = Form(""),
    amount: float = Form(...),
    mad_category_id: int = Form(...),
    expense_category_id: int = Form(...),
    payment_mode_id: int = Form(...),
    referred_by: str = Form(""),
    notes: str = Form("")
):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    db = get_db()
    try:
        expense_date = datetime.strptime(date, "%Y-%m-%d").date()
        expense = Expense(
            voucher_no=voucher_no,
            date=expense_date,
            name=name,
            phone=phone,
            amount=amount,
            mad_category_id=mad_category_id,
            expense_category_id=expense_category_id,
            payment_mode_id=payment_mode_id,
            referred_by=referred_by,
            notes=notes
        )
        db.add(expense)
        
        # Check if person exists as Beneficiary or Staff, if not create as Beneficiary
        existing_person = db.query(Person).filter(func.lower(Person.name) == name.lower(), Person.person_type.in_(['Beneficiary', 'Staff'])).first()
        if not existing_person:
            new_person = Person(name=name, phone=phone, person_type='Beneficiary')
            db.add(new_person)
            
        db.commit()
        return RedirectResponse(url="/expenses", status_code=303)
    except Exception as e:
        db.rollback()
        template = env.get_template("add_voucher.html")
        return HTMLResponse(content=template.render(
            request=request,
            error=f"Error adding voucher: {str(e)}",
            mad_categories=db.query(MadCategory).all(),
            expense_categories=db.query(ExpenseCategory).all(),
            payment_modes=db.query(PaymentMode).all(),
            voucher_no=voucher_no,
            date=date,
            name=name,
            phone=phone,
            amount=amount,
            mad_category_id=mad_category_id,
            expense_category_id=expense_category_id,
            payment_mode_id=payment_mode_id,
            referred_by=referred_by,
            notes=notes
        ))
    finally:
        db.close()

@app.get("/edit-voucher/{expense_id}")
def edit_voucher_page(request: Request, expense_id: int):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    db = get_db()
    try:
        expense = db.query(Expense).filter(Expense.id == expense_id).first()
        if not expense:
            return RedirectResponse(url="/expenses", status_code=303)
        
        template = env.get_template("edit_voucher.html")
        return HTMLResponse(content=template.render(
            request=request,
            expense=expense,
            mad_categories=db.query(MadCategory).all(),
            expense_categories=db.query(ExpenseCategory).all(),
            payment_modes=db.query(PaymentMode).all()
        ))
    finally:
        db.close()

@app.post("/edit-voucher/{expense_id}")
async def edit_voucher(
    request: Request,
    expense_id: int,
    voucher_no: str = Form(...),
    date: str = Form(...),
    name: str = Form(...),
    phone: str = Form(""),
    amount: float = Form(...),
    mad_category_id: int = Form(...),
    expense_category_id: int = Form(...),
    payment_mode_id: int = Form(...),
    referred_by: str = Form(""),
    notes: str = Form("")
):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    db = get_db()
    try:
        expense = db.query(Expense).filter(Expense.id == expense_id).first()
        if expense:
            expense.date = datetime.strptime(date, "%Y-%m-%d").date()
            expense.voucher_no = voucher_no
            expense.name = name
            expense.phone = phone
            expense.amount = amount
            expense.mad_category_id = mad_category_id
            expense.expense_category_id = expense_category_id
            expense.payment_mode_id = payment_mode_id
            expense.referred_by = referred_by
            expense.notes = notes
            db.commit()
        return RedirectResponse(url="/expenses", status_code=303)
    except Exception as e:
        db.rollback()
        expense = db.query(Expense).filter(Expense.id == expense_id).first()
        template = env.get_template("edit_voucher.html")
        return HTMLResponse(content=template.render(
            request=request,
            expense=expense,
            error=f"Error updating voucher: {str(e)}",
            mad_categories=db.query(MadCategory).all(),
            expense_categories=db.query(ExpenseCategory).all(),
            payment_modes=db.query(PaymentMode).all()
        ))
    finally:
        db.close()

@app.get("/delete-voucher/{expense_id}")
def delete_voucher_page(request: Request, expense_id: int):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    db = get_db()
    try:
        expense = db.query(Expense).filter(Expense.id == expense_id).first()
        if not expense:
            return RedirectResponse(url="/expenses", status_code=303)
        
        template = env.get_template("delete_voucher_confirm.html")
        return HTMLResponse(content=template.render(
            request=request,
            expense=expense
        ))
    finally:
        db.close()

@app.post("/delete-voucher/{expense_id}")
def delete_voucher(request: Request, expense_id: int, password: str = Form(...)):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    
    if password != DEFAULT_ADMIN_PASSWORD:
        db = get_db()
        try:
            expense = db.query(Expense).filter(Expense.id == expense_id).first()
            template = env.get_template("delete_voucher_confirm.html")
            return HTMLResponse(content=template.render(
                request=request,
                expense=expense,
                error="Incorrect password. Deletion cancelled."
            ))
        finally:
            db.close()
            
    db = get_db()
    try:
        expense = db.query(Expense).filter(Expense.id == expense_id).first()
        if expense:
            db.delete(expense)
            db.commit()
        return RedirectResponse(url="/expenses", status_code=303)
    finally:
        db.close()
@app.get("/settings/opening-balances")
def opening_balances_settings(request: Request):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    db = get_db()
    try:
        mad_categories = db.query(MadCategory).all()
        registers = db.query(Register).all()
        payment_modes = db.query(PaymentMode).all()
        
        # Fetch existing MAD opening balances mapped as string keys "madId_modeId"
        initial_balances = db.query(MadInitialBalance).all()
        existing_balances = {f"{ib.mad_category_id}_{ib.payment_mode_id}": ib.amount for ib in initial_balances}
        
        template = env.get_template("opening_balances.html")
        return HTMLResponse(content=template.render(
            request=request,
            mad_categories=mad_categories,
            registers=registers,
            payment_modes=payment_modes,
            existing_balances=existing_balances
        ))
    finally:
        db.close()

@app.post("/update-initial-balances")
async def update_initial_balances(request: Request):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    
    db = get_db()
    try:
        form = await request.form()
        confirm_password = form.get("confirm_password")
        
        # Verify admin password
        if confirm_password != DEFAULT_ADMIN_PASSWORD:
            mad_categories = db.query(MadCategory).all()
            registers = db.query(Register).all()
            payment_modes = db.query(PaymentMode).all()
            
            # Extract form values to preserve them in the UI
            entered_balances = {}
            for key, value in form.items():
                if key.startswith("balance_"):
                    try:
                        parts = key.split("_")
                        mad_id = int(parts[1])
                        mode_id = int(parts[2])
                        amount = float(value or 0.0)
                        entered_balances[f"{mad_id}_{mode_id}"] = amount
                    except Exception:
                        continue
            
            template = env.get_template("opening_balances.html")
            return HTMLResponse(content=template.render(
                request=request,
                mad_categories=mad_categories,
                registers=registers,
                payment_modes=payment_modes,
                existing_balances=entered_balances,
                error="Incorrect admin password. Balance allocation was not saved!"
            ))
            
        # If password is correct, save the new balances
        saved_details = []
        for key, value in form.items():
            if key.startswith("balance_"):
                try:
                    parts = key.split("_")
                    mad_id = int(parts[1])
                    mode_id = int(parts[2])
                    amount = float(value or 0)
                    
                    # Update or create the MadInitialBalance record
                    ib = db.query(MadInitialBalance).filter(
                        and_(MadInitialBalance.mad_category_id == mad_id, MadInitialBalance.payment_mode_id == mode_id)
                    ).first()
                    
                    if not ib:
                        ib = MadInitialBalance(mad_category_id=mad_id, payment_mode_id=mode_id)
                        db.add(ib)
                    
                    ib.amount = amount
                    
                    if amount > 0:
                        mad = db.query(MadCategory).filter(MadCategory.id == mad_id).first()
                        mode = db.query(PaymentMode).filter(PaymentMode.id == mode_id).first()
                        saved_details.append({
                            "mad": mad.name if mad else f"ID {mad_id}",
                            "mode": mode.name if mode else f"ID {mode_id}",
                            "amount": amount
                        })
                except (ValueError, IndexError):
                    continue
        
        db.commit()
        
        # Recalculate and cache payment mode initial balances
        payment_modes = db.query(PaymentMode).all()
        for mode in payment_modes:
            total_initial = db.query(func.sum(MadInitialBalance.amount)).filter(
                MadInitialBalance.payment_mode_id == mode.id
            ).scalar() or 0.0
            mode.initial_balance = total_initial
            
        # Recalculate and cache MAD initial balances
        mad_categories = db.query(MadCategory).all()
        for mad in mad_categories:
            total_initial = db.query(func.sum(MadInitialBalance.amount)).filter(
                MadInitialBalance.mad_category_id == mad.id
            ).scalar() or 0.0
            mad.initial_balance = total_initial
            
        db.commit()
        
        # Re-fetch the saved balances from DB to populate existing_balances
        initial_balances = db.query(MadInitialBalance).all()
        existing_balances = {f"{ib.mad_category_id}_{ib.payment_mode_id}": ib.amount for ib in initial_balances}
        registers = db.query(Register).all()
        
        template = env.get_template("opening_balances.html")
        return HTMLResponse(content=template.render(
            request=request,
            mad_categories=mad_categories,
            registers=registers,
            payment_modes=payment_modes,
            existing_balances=existing_balances,
            success=True,
            saved_details=saved_details
        ))
    finally:
        db.close()

@app.get("/directory")
def directory_page(request: Request, search: str = None, person_type: str = None):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    
    db = get_db()
    try:
        query = db.query(Person)
        if search:
            query = query.filter(Person.name.ilike(f"%{search}%") | Person.phone.ilike(f"%{search}%") | Person.address.ilike(f"%{search}%"))
        if person_type:
            query = query.filter(Person.person_type == person_type)
            
        people = query.order_by(Person.name).all()
        template = env.get_template("directory.html")
        return HTMLResponse(content=template.render(request=request, people=people, search=search, person_type=person_type))
    finally:
        db.close()

@app.post("/directory/add")
def add_directory_profile(
    request: Request,
    name: str = Form(...),
    phone: str = Form(""),
    address: str = Form(""),
    person_type: str = Form(...)
):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    
    db = get_db()
    try:
        new_person = Person(name=name, phone=phone, address=address, person_type=person_type)
        db.add(new_person)
        db.commit()
        return RedirectResponse(url="/directory", status_code=303)
    finally:
        db.close()

@app.post("/directory/edit")
def edit_directory_profile(
    request: Request,
    person_id: int = Form(...),
    name: str = Form(...),
    phone: str = Form(""),
    address: str = Form(""),
    person_type: str = Form(...)
):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    
    db = get_db()
    try:
        person = db.query(Person).filter(Person.id == person_id).first()
        if person:
            old_name = person.name
            
            person.name = name
            person.phone = phone
            person.address = address
            person.person_type = person_type
            
            # Update all related receipts and expenses with new details
            from sqlalchemy import func
            db.query(Receipt).filter(func.lower(Receipt.name) == old_name.lower()).update(
                {"name": name, "phone": phone, "address": address}, 
                synchronize_session=False
            )
            db.query(Expense).filter(func.lower(Expense.name) == old_name.lower()).update(
                {"name": name, "phone": phone}, 
                synchronize_session=False
            )
            
            db.commit()
        return RedirectResponse(url="/directory", status_code=303)
    finally:
        db.close()

@app.post("/directory/delete/{person_id}")
def delete_directory_profile(request: Request, person_id: int):
    auth_redirect = require_auth(request)
    if auth_redirect: return auth_redirect
    
    db = get_db()
    try:
        person = db.query(Person).filter(Person.id == person_id).first()
        if person:
            db.delete(person)
            db.commit()
        return RedirectResponse(url="/directory", status_code=303)
    finally:
        db.close()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
